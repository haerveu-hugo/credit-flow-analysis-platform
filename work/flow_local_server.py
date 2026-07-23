from __future__ import annotations

import cgi
import hashlib
import html
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import traceback
import uuid
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, quote, urlparse

ROOT = Path(__file__).resolve().parents[1]
RUNTIME = Path(os.environ.get("PLATFORM_RUNTIME", ROOT / "runtime" / "dependencies"))
PYTHON = Path(os.environ.get("PYTHON_BIN", RUNTIME / "python" / "bin" / "python3"))
POPPLER = Path(os.environ.get("POPPLER_BIN", RUNTIME / "bin"))
UPLOADS = ROOT / "tmp" / "flow_uploads"
RESULTS = ROOT / "tmp" / "flow_results"
CACHE = ROOT / "tmp" / "flow_cache"
FLOW_OCR_CACHE_VERSION = "flow-ocr-v4-fullpage-progress"
JOBS: dict[str, dict] = {}
JOBS_LOCK = threading.Lock()

sys.path.insert(0, str(ROOT / "work"))
sys.path.insert(0, str(ROOT / "work" / "ocr_deps"))

import pandas as pd
import pdfplumber
from ocr_credit_pdf import ocr_images, prepare_and_split, render_pdf


IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"}
EXCEL_EXTS = {".xlsx", ".xls", ".csv"}
PDF_EXTS = {".pdf"}
DEFAULT_EXCLUDE_KEYWORDS = [
    "手续费",
    "账户管理费",
    "短信费",
    "利息",
    "结息",
    "冲销",
    "贷款发放",
    "贷款入账",
    "放贷款",
    "借款",
    "理财赎回",
    "基金赎回",
    "银证转入",
    "保证金退回",
    "微信内部流水",
]


class PasswordRequiredError(Exception):
    def __init__(self, filename: str, message: str | None = None):
        self.filename = filename
        super().__init__(message or f"{filename} 需要输入文件密码。")


class JobCancelledError(Exception):
    pass


def set_job(job_id: str, **updates) -> None:
    with JOBS_LOCK:
        job = JOBS.setdefault(job_id, {})
        job.update(updates)


def get_job(job_id: str) -> dict:
    with JOBS_LOCK:
        return dict(JOBS.get(job_id) or {})


def is_job_cancelled(job_id: str) -> bool:
    job = get_job(job_id)
    return bool(job.get("cancel_requested")) or job.get("status") in {"cancelling", "cancelled"}


def raise_if_cancelled(should_cancel=None) -> None:
    if should_cancel and should_cancel():
        raise JobCancelledError("已停止统计")


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_passwords(raw) -> list[str]:
    if raw is None:
        return [""]
    text = str(raw).strip()
    values = []
    for item in re.split(r"[,，;；\n\r]+", text):
        item = normalize_text(item)
        if not item:
            continue
        if re.search(r"[=:：]", item):
            tail = re.split(r"[=:：]", item, maxsplit=1)[-1].strip()
            if tail:
                item = tail
        keyword_match = re.search(r"(?:密码|口令|password|pass)\s*([A-Za-z0-9]{3,32})", item, re.I)
        if keyword_match:
            values.append(keyword_match.group(1))
            continue
        if re.fullmatch(r"[A-Za-z0-9@#$%._-]{3,64}", item):
            values.append(item)
            continue
        code_candidates = re.findall(r"(?<!\d)[A-Za-z0-9]{4,16}(?!\d)", item)
        values.extend(code_candidates)
    unique = []
    for value in values:
        if value not in unique:
            unique.append(value)
    return ["", *unique] if "" not in unique else unique


def cache_key(path: Path, password: str | None = None) -> str:
    digest = hashlib.sha256()
    digest.update(FLOW_OCR_CACHE_VERSION.encode("utf-8"))
    digest.update(file_hash(path).encode("utf-8"))
    if password:
        digest.update(b":password:")
        digest.update(password.encode("utf-8"))
    return digest.hexdigest()


def money_to_float(value) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null", "--", "-"}:
        return None
    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]
    text = (
        text.replace(",", "")
        .replace("，", "")
        .replace("￥", "")
        .replace("¥", "")
        .replace("人民币", "")
        .replace("元", "")
        .replace(" ", "")
    )
    text = re.sub(r"[^\d.+-]", "", text)
    if not text or text in {".", "+", "-"}:
        return None
    try:
        amount = float(Decimal(text))
    except (InvalidOperation, ValueError):
        return None
    return -abs(amount) if negative else amount


def money(value: float | int | None) -> str:
    amount = float(value or 0)
    return f"{amount:,.2f}".rstrip("0").rstrip(".")


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value)
    if text.lower() == "nan":
        return ""
    return re.sub(r"\s+", " ", text).strip()


def parse_date(value, default_year: int | None = None) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime("%Y-%m-%d")
    text = normalize_text(value)
    if not text:
        return None
    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    patterns = [
        r"(20\d{2}|19\d{2})[-/.](\d{1,2})[-/.](\d{1,2})",
        r"(20\d{2}|19\d{2})(\d{2})(\d{2})",
        r"(\d{1,2})[-/.](\d{1,2})",
    ]
    for idx, pattern in enumerate(patterns):
        match = re.search(pattern, text)
        if not match:
            continue
        if idx == 2:
            year = default_year or datetime.now().year
            month, day = int(match.group(1)), int(match.group(2))
        else:
            year, month, day = int(match.group(1)), int(match.group(2)), int(match.group(3))
        try:
            return datetime(year, month, day).strftime("%Y-%m-%d")
        except ValueError:
            return None
    return None


def parse_short_bank_date(value) -> str | None:
    text = normalize_text(value)
    match = re.fullmatch(r"(\d{2})(\d{2})(\d{2})", text)
    if not match:
        return parse_date(text)
    year = 2000 + int(match.group(1))
    month = int(match.group(2))
    day = int(match.group(3))
    try:
        return datetime(year, month, day).strftime("%Y-%m-%d")
    except ValueError:
        return None


def month_key(date_text: str | None) -> str:
    return (date_text or "未识别日期")[:7] if date_text else "未识别日期"


def detect_direction(text: str, amount: float | None = None) -> str:
    content = normalize_text(text)
    if content in {"贷", "贷方"}:
        return "收入"
    if content in {"借", "借方"}:
        return "支出"
    if amount is not None and amount < 0:
        return "支出"
    if re.search(r"收入|贷方|进账|入账|存入|转入|收款|来账|贷记|代发", content):
        return "收入"
    if re.search(r"支出|借方|出账|取现|转出|付款|消费|扣款|借记|还款|还贷|手续费|服务费|年费", content):
        return "支出"
    return "收入" if amount and amount > 0 else "未知"


def infer_counterparty(parts: list[str]) -> str:
    candidates = []
    skip = re.compile(r"日期|时间|余额|金额|收入|支出|借方|贷方|摘要|用途|备注|交易|人民币|账号|卡号|流水|页|合计")
    for part in parts:
        text = normalize_text(part)
        if not text or skip.search(text):
            continue
        if parse_date(text):
            continue
        if money_to_float(text) is not None:
            continue
        if len(text) >= 2:
            candidates.append(text)
    return candidates[0] if candidates else "未识别"


def normalize_account(value, fallback: str = "") -> str:
    text = normalize_text(value)
    if text:
        digits = re.sub(r"[^0-9*]", "", text)
        if len(digits) >= 6:
            return digits
        if len(text) >= 2 and not re.search(r"账号|账户|卡号|户号", text):
            return text
    return fallback or "未识别账户"


def infer_account_from_text(text: str, fallback: str = "") -> str:
    wechat_match = re.search(r"兹证明[:：]\s*([\u4e00-\u9fffA-Za-z0-9·]{2,20}).*?微信号[:：]\s*([A-Za-z0-9_*-]{4,})", text, re.S)
    if wechat_match:
        name = normalize_text(wechat_match.group(1))
        wechat_id = normalize_text(wechat_match.group(2))
        suffix = re.sub(r"[^A-Za-z0-9]", "", wechat_id)[-4:] or wechat_id[-4:]
        return f"{name}微信({suffix})"
    for pattern in [
        r"(?:本方账号|交易账号|账\s*号|账户(?:号码|账号|号)?|卡\s*号|户号)[:： \t]*([0-9* ]{6,})",
        r"(?<!对方)(?:本方户名|账户名称|户\s*名)[:： \t]*([\u4e00-\u9fffA-Za-z0-9（）()·]{2,})",
    ]:
        match = re.search(pattern, text)
        if match:
            return normalize_account(match.group(1), fallback)
    return fallback or "未识别账户"


def infer_account_from_filename(filename: str) -> str:
    text = normalize_text(filename)
    for pattern in [
        r"[，,]\s*([0-9*]{8,})\s*[，,]\s*人民币",
        r"(?:账号|账户|户号)[：:\s]*([0-9*]{8,})",
    ]:
        match = re.search(pattern, text)
        if match:
            return normalize_account(match.group(1), "")
    return ""


def normalize_source_account(source: str) -> str:
    if "微信支付交易明细" in source:
        return "微信支付"
    if "支付宝交易明细" in source:
        return "支付宝"
    filename = source.split(" / ", 1)[0] if source else ""
    stem = Path(filename).stem if filename else ""
    stem = re.sub(r"[（(]?20\d{6,8}\s*[-—~～至到_]\s*20\d{6,8}[）)]?", "", stem)
    stem = re.sub(r"20\d{2}[./-]\d{1,2}(?:[./-]\d{1,2})?\s*[-—~～至到_]\s*20\d{2}[./-]\d{1,2}(?:[./-]\d{1,2})?", "", stem)
    stem = re.sub(r"[-_（）()\s]+$", "", stem).strip()
    return stem or "未识别账户"


def extract_statement_info(text: str, filename: str = "") -> dict:
    content = normalize_text(text)
    info = {"filename": filename}
    patterns = {
        "户名": r"(?:客户姓名|(?<!对方)户\s*名|(?<!对方)账户名称|客户名称)[:： \t]*([\u4e00-\u9fffA-Za-z0-9（）()·]{2,30})",
        "账号": r"(?:卡\s*号|账\s*号|账户号码|账户)[:： \t]*([0-9* ]{6,})",
        "币种": r"币\s*种[:：]\s*([\u4e00-\u9fffA-Za-z]{2,10})",
        "起止日期": r"起止日期[:：\s]*([0-9]{4}[-/.]?[0-9]{1,2}[-/.]?[0-9]{1,2}\s*(?:至|到|一|-|—|~|～)+\s*[0-9]{4}[-/.]?[0-9]{1,2}[-/.]?[0-9]{1,2})",
        "电子流水号": r"电子流水号[:：\s]*([A-Za-z0-9-]{8,})",
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, content)
        if match:
            info[key] = normalize_text(match.group(1))
    company_match = re.search(r"用户所属公司[:：]\s*(.+?)(?:\s+打印时间[:：]|$)", content)
    if company_match:
        info["户名"] = normalize_text(company_match.group(1))
    filename_account = infer_account_from_filename(filename)
    if filename_account:
        info["账号"] = filename_account
    filename_dates = re.search(r"(20\d{6,8})[-—~～至到_](20\d{6,8})", normalize_text(filename))
    if filename_dates and "起止日期" not in info:
        info["起止日期"] = f"{filename_dates.group(1)} 至 {filename_dates.group(2)}"
    start_match = re.search(r"开始日期[:：\s]*(20\d{2}[/.-]\d{1,2}[/.-]\d{1,2})", content)
    end_match = re.search(r"结束日期[:：\s]*(20\d{2}[/.-]\d{1,2}[/.-]\d{1,2})", content)
    if start_match and end_match:
        info["起止日期"] = f"{start_match.group(1)} 至 {end_match.group(1)}"
    if "微信支付交易明细证明" in content:
        wechat_identity = re.search(
            r"兹证明[:：]\s*([\u4e00-\u9fffA-Za-z0-9·]{2,20}).*?微信号[:：]\s*([A-Za-z0-9_*-]{4,})",
            content,
            re.S,
        )
        if wechat_identity:
            info["户名"] = normalize_text(wechat_identity.group(1))
            info["账号"] = infer_account_from_text(content, "")
    return {k: v for k, v in info.items() if v}


def infer_account_from_df(df: pd.DataFrame, fallback: str) -> str:
    for row_idx in range(min(len(df), 35)):
        joined = " ".join(normalize_text(v) for v in df.iloc[row_idx].tolist())
        account = infer_account_from_text(joined, "")
        if account and account != "未识别账户":
            return account
    return fallback


TABLE_HEADER_RE = re.compile(
    r"交易日期.*(?:交易金额|收入|支出|余额|对方|对手)|"
    r"交易时间.*(?:交易摘要|交易金额|本次余额|对手信息)|"
    r"收入[/／》]?支出金额|贷方发生额|借方发生额"
)
TABLE_FOOTER_RE = re.compile(
    r"该交易明细|明细内容仅供参考|第\s*\d+\s*页|共\s*\d+\s*页|"
    r"制表|打印|操作员|经办|复核|温馨提示|说明[:：]"
)


def is_table_header(line: str) -> bool:
    return bool(TABLE_HEADER_RE.search(normalize_text(line)))


def is_table_footer(line: str) -> bool:
    return bool(TABLE_FOOTER_RE.search(normalize_text(line)) or is_summary_text(line))


SUMMARY_LINE_RE = re.compile(
    r"合计|汇总|小计|总计|累计|本页合计|本月合计|本年合计|收入合计|支出合计|"
    r"借方合计|贷方合计|发生额合计|交易笔数|笔数合计|总收入|总支出|总金额|"
    r"流水合计|收入汇总|支出汇总|汇总流水"
)


def is_summary_text(text: str) -> bool:
    content = normalize_text(text)
    return bool(content and SUMMARY_LINE_RE.search(content))


def is_summary_row(row: list[str]) -> bool:
    return is_summary_text(" ".join(normalize_text(v) for v in row))


def make_txn(
    *,
    date: str | None,
    account: str,
    counterparty: str,
    summary: str,
    income: float | None,
    expense: float | None,
    amount: float | None,
    balance: float | None,
    source: str,
    raw_key: str | None = None,
    preserve_signed_columns: bool = False,
) -> dict | None:
    if preserve_signed_columns:
        income = income if income is not None and income != 0 else None
        expense = expense if expense is not None and expense != 0 else None
    else:
        income = income if income and income > 0 else None
        expense = expense if expense and expense > 0 else None
    if amount is not None and not income and not expense:
        if abs(float(amount or 0)) < 0.01:
            return None
        if amount < 0:
            expense = abs(amount)
        else:
            direction = detect_direction(f"{summary} {counterparty}", amount)
            if direction == "支出":
                expense = abs(amount)
            else:
                income = abs(amount)
    if not date or (income is None and expense is None):
        return None
    direction = "收入" if income is not None else "支出"
    text = f"{counterparty} {summary}"
    inferred_interest = (
        income is not None
        and float(income or 0) <= 1000
        and not expense
        and not normalize_text(summary)
        and bool(re.search(r"银行", counterparty or ""))
        and (date[-2:] in {"20", "21", "22"})
    )
    category = "银行结息" if re.search(r"季度结息|结息|利息", text) or inferred_interest else "普通流水"
    default_excluded = category == "银行结息" or any(keyword in text for keyword in DEFAULT_EXCLUDE_KEYWORDS)
    return {
        "date": date,
        "month": month_key(date),
        "account": normalize_account(account, normalize_source_account(source)),
        "counterparty": counterparty or "未识别",
        "summary": summary or "",
        "income": round(float(income or 0), 2),
        "expense": round(float(expense or 0), 2),
        "amount": round(float(income or 0) - float(expense or 0), 2),
        "balance": round(float(balance or 0), 2) if balance is not None else None,
        "direction": direction,
        "category": category,
        "default_excluded": default_excluded,
        "exclude_reason": "银行结息" if category == "银行结息" else ("常见非经营流水" if default_excluded else ""),
        "source": source,
        "raw_key": raw_key or "",
    }


def find_header_map(df: pd.DataFrame) -> tuple[int | None, dict[str, int]]:
    best_row = None
    best_score = 0
    best_map: dict[str, int] = {}
    for row_idx in range(min(len(df), 40)):
        row = [normalize_text(v) for v in df.iloc[row_idx].tolist()]
        score = 0
        mapping: dict[str, int] = {}
        for idx, cell in enumerate(row):
            compact = re.sub(r"\s+", "", cell)
            if re.search(r"交易日期|记账日期|记账日|会计日期|起息日|日期|交易时间|发生日期", compact):
                mapping.setdefault("date", idx)
                score += 3
            if re.search(r"收入[/／]支出金额|收[/／]支金额|收入支出金额|借贷金额", compact):
                mapping.setdefault("amount", idx)
                score += 4
                continue
            if re.search(r"余额|账户余额|本次余额", compact):
                mapping.setdefault("balance", idx)
                score += 2
                continue
            if re.search(r"对方户名|对方单位|交易对手|对方名称|对方账号名称|收付款人|对手户名|对手名称", compact):
                mapping["counterparty"] = idx
                score += 4
                continue
            if re.search(r"户名|客户名称", compact) and not re.search(r"本方|我方|账户名称|客户名称[:：]?$", compact):
                mapping.setdefault("counterparty", idx)
                score += 2
                continue
            if (
                re.search(r"本方账号|交易账号|账号|账户号|卡号|户号|账户", compact)
                and not re.search(r"对方|对手|余额|序号|子账号", compact)
            ):
                mapping.setdefault("account", idx)
                score += 2
            if re.search(r"摘要|用途|附言|备注|交易类型|业务摘要|说明", compact):
                mapping.setdefault("summary", idx)
                score += 2
            if re.search(r"收入|贷方发生额|贷方金额|进账|存入|转入金额", compact):
                mapping.setdefault("income", idx)
                score += 3
            if re.search(r"支出|借方发生额|借方金额|出账|取出|转出金额", compact):
                mapping.setdefault("expense", idx)
                score += 3
            if re.search(r"发生额|交易金额|金额", compact) and not re.search(r"余额|收入|支出|贷方|借方|转入|转出", compact):
                mapping.setdefault("amount", idx)
                score += 2
            if re.search(r"收[/／]?支|收支|借贷|方向", compact):
                mapping.setdefault("direction", idx)
                score += 1
        if score > best_score and "date" in mapping and (
            "income" in mapping or "expense" in mapping or "amount" in mapping
        ):
            best_row, best_score, best_map = row_idx, score, mapping
    return best_row, best_map


def extract_excel(path: Path, progress=None, passwords: list[str] | None = None, should_cancel=None) -> list[dict]:
    raise_if_cancelled(should_cancel)
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            sheets = {"CSV": pd.read_csv(path, header=None, dtype=str, encoding_errors="ignore")}
        elif suffix == ".xls":
            try:
                sheets = pd.read_excel(path, sheet_name=None, header=None, dtype=str, engine="xlrd")
            except Exception:
                converted = RESULTS / f"{path.stem}-converted"
                converted.mkdir(parents=True, exist_ok=True)
                soffice = POPPLER / "soffice"
                profile = converted / "lo-profile"
                profile.mkdir(parents=True, exist_ok=True)
                if soffice.exists():
                    subprocess.run(
                        [
                            str(soffice),
                            f"-env:UserInstallation={profile.resolve().as_uri()}",
                            "--headless",
                            "--convert-to",
                            "xlsx",
                            "--outdir",
                            str(converted),
                            str(path),
                        ],
                        check=False,
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        timeout=120,
                    )
                matches = list(converted.glob("*.xlsx"))
                if not matches:
                    raise ValueError("无法读取该 .xls 文件；若文件有密码，请先另存为无密码的 .xlsx。")
                sheets = pd.read_excel(matches[0], sheet_name=None, header=None, dtype=str)
        else:
            sheets = pd.read_excel(path, sheet_name=None, header=None, dtype=str)
    except Exception as exc:
        message = str(exc)
        if suffix in {".xlsx", ".xls"} and re.search(r"password|encrypted|加密|密码|not a zip file|File is not a zip", message, re.I):
            raise ValueError(
                f"{path.name} 可能是加密 Excel。当前环境不能直接解密 Excel，请先在 Excel/WPS 中输入密码打开，再另存为无密码的 xlsx 后上传。",
            ) from exc
        raise

    txns: list[dict] = []
    total_sheets = len(sheets) or 1
    for sheet_no, (sheet_name, df) in enumerate(sheets.items(), start=1):
        raise_if_cancelled(should_cancel)
        if progress:
            progress(int(20 + 50 * sheet_no / total_sheets), f"读取 Excel 工作表 {sheet_no}/{total_sheets}")
        header_row, mapping = find_header_map(df)
        if mapping.get("income") == mapping.get("expense") and mapping.get("income") is not None:
            mapping.setdefault("amount", mapping["income"])
            mapping.pop("income", None)
            mapping.pop("expense", None)
        start = (header_row + 1) if header_row is not None else 0
        default_year = None
        sheet_account = infer_account_from_df(df, normalize_source_account(path.name))
        header_cells = [normalize_text(v) for v in df.iloc[header_row].tolist()] if header_row is not None else []
        signed_amount_column = False
        if "amount" in mapping:
            amount_col = mapping["amount"]
            sample_amounts = [
                money_to_float(df.iloc[row_idx, amount_col])
                for row_idx in range(start, min(len(df), start + 300))
            ]
            signed_amount_column = any(value is not None and value < 0 for value in sample_amounts)
        income_nonzero = sum(
            1
            for row_idx in range(start, min(len(df), start + 500))
            if "income" in mapping and (money_to_float(normalize_text(df.iloc[row_idx, mapping["income"]])) or 0) != 0
        )
        expense_nonzero = sum(
            1
            for row_idx in range(start, min(len(df), start + 500))
            if "expense" in mapping and (money_to_float(normalize_text(df.iloc[row_idx, mapping["expense"]])) or 0) != 0
        )
        single_direction_column = (
            "income" in mapping
            and "expense" in mapping
            and (
                (income_nonzero > 0 and income_nonzero >= max(1, expense_nonzero) * 10)
                or (expense_nonzero > 0 and expense_nonzero >= max(1, income_nonzero) * 10)
            )
        )
        date_samples = []
        if "date" in mapping:
            for row_idx in range(start, min(len(df), start + 80)):
                sample_date = parse_date(df.iloc[row_idx, mapping["date"]])
                if sample_date:
                    date_samples.append(sample_date)
        ascending_rows = len(date_samples) >= 2 and date_samples[-1] >= date_samples[0]
        previous_balance = None
        summary_cols = [
            idx for idx, cell in enumerate(header_cells)
            if re.search(r"摘要|用途|附言|备注|交易类型|业务摘要|说明", re.sub(r"\s+", "", cell))
        ]
        for row_idx in range(start, len(df)):
            if row_idx % 50 == 0:
                raise_if_cancelled(should_cancel)
            row = [normalize_text(v) for v in df.iloc[row_idx].tolist()]
            if is_summary_row(row):
                continue
            joined = " ".join(row)
            if "date" in mapping:
                date_value = row[mapping["date"]]
                date = parse_date(date_value, default_year) or parse_short_bank_date(date_value)
            else:
                date = parse_date(joined, default_year)
            if date:
                default_year = int(date[:4])
            income = money_to_float(row[mapping["income"]]) if "income" in mapping and mapping["income"] < len(row) else None
            expense = money_to_float(row[mapping["expense"]]) if "expense" in mapping and mapping["expense"] < len(row) else None
            amount = money_to_float(row[mapping["amount"]]) if "amount" in mapping and mapping["amount"] < len(row) else None
            balance = money_to_float(row[mapping["balance"]]) if "balance" in mapping and mapping["balance"] < len(row) else None
            account = row[mapping["account"]] if "account" in mapping and mapping["account"] < len(row) else sheet_account
            account_digits = re.sub(r"\D", "", account or "")
            if account_digits and len(account_digits) < 6:
                account = sheet_account
            counterparty = row[mapping["counterparty"]] if "counterparty" in mapping and mapping["counterparty"] < len(row) else ""
            counterparty = counterparty or infer_counterparty(row)
            if summary_cols:
                summary = " ".join(row[idx] for idx in summary_cols if idx < len(row) and row[idx])
            else:
                summary = row[mapping["summary"]] if "summary" in mapping and mapping["summary"] < len(row) else joined
            if signed_amount_column and amount is not None and "direction" not in mapping and income is None and expense is None:
                if amount > 0:
                    income, expense = amount, None
                elif amount < 0:
                    income, expense = None, abs(amount)
            if "direction" in mapping and mapping["direction"] < len(row) and amount is not None:
                direction_text = row[mapping["direction"]]
                direction = detect_direction(direction_text, amount)
                if direction == "收入":
                    income, expense = abs(amount), None
                elif direction == "支出":
                    income, expense = None, abs(amount)
            else:
                direction_text = ""
            if single_direction_column:
                single_amount = abs(float(income or expense or 0))
                if ascending_rows and previous_balance is not None and balance is not None and single_amount:
                    effect = float(balance) - float(previous_balance)
                    tolerance = max(0.05, single_amount * 0.0001)
                    if abs(abs(effect) - single_amount) <= tolerance:
                        if effect > 0:
                            income, expense = single_amount, None
                        elif effect < 0:
                            income, expense = None, single_amount
                elif single_amount:
                    inferred = detect_direction(summary, None)
                    if inferred == "支出":
                        income, expense = None, single_amount
                    elif inferred == "收入":
                        income, expense = single_amount, None
            is_wechat_internal = "微信支付交易明细" in path.name and "其他" in normalize_text(direction_text)
            if "不计" in normalize_text(direction_text):
                continue
            txn = make_txn(
                date=date,
                account=account,
                counterparty=counterparty,
                summary=summary,
                income=income,
                expense=expense,
                amount=amount,
                balance=balance,
                source=f"{path.name} / {sheet_name}",
            )
            if txn:
                if is_wechat_internal:
                    txn["category"] = "微信内部流水"
                    txn["default_excluded"] = True
                    txn["exclude_reason"] = "微信内部流水"
                    txn["direction"] = "其他"
                    txn["other_amount"] = round(abs(float(amount or income or expense or 0)), 2)
                    txn["income"] = 0.0
                    txn["expense"] = 0.0
                    txn["amount"] = 0.0
                txns.append(txn)
            if balance is not None:
                previous_balance = balance
    return txns


def extract_pdf_text(path: Path, passwords: list[str] | None = None, should_cancel=None, progress=None) -> tuple[str, str]:
    raise_if_cancelled(should_cancel)
    passwords = passwords or [""]
    last_error = ""
    for password in passwords:
        chunks = []
        try:
            with pdfplumber.open(path, password=password or None) as pdf:
                for i, page in enumerate(pdf.pages, start=1):
                    raise_if_cancelled(should_cancel)
                    if progress:
                        progress(10 + int(65 * (i - 1) / max(len(pdf.pages), 1)), f"读取 PDF 第 {i}/{len(pdf.pages)} 页")
                    text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
                    chunks.append(f"【第 {i} 页】\n{text}")
                    for table in page.extract_tables() or []:
                        for row in table:
                            chunks.append(" | ".join(normalize_text(cell) for cell in row or []))
            return "\n".join(chunks), password
        except Exception as exc:
            last_error = str(exc)
            error_text = f"{type(exc).__name__}: {last_error}"
            if (not password and len(passwords) > 1) or re.search(r"password|encrypted|加密|密码|PdfminerException", error_text, re.I):
                continue
            if not re.search(r"password|encrypted|加密|密码", error_text, re.I):
                return "", password
    if len(passwords) <= 1 and not passwords[0]:
        raise PasswordRequiredError(path.name)
    raise PasswordRequiredError(path.name, f"{path.name} 密码不正确，或还需要输入正确密码。")


def is_corporate_query_statement(text: str, filename: str = "") -> bool:
    content = normalize_text(text)
    name = normalize_text(filename)
    return bool(
        "交易查询" in name
        and "交易日期" in content
        and "借方(出账)" in content
        and "贷方(入账)" in content
        and "收(付)方名称" in content
    )


def is_abc_account_detail_statement(text: str, filename: str = "") -> bool:
    content = normalize_text(text)
    return bool(
        "账户明细" in content
        and "交易时间 收入金额 支出金额 账户余额 对方账号 对方户名 交易用途 对方开户行 摘要" in content
    )


def join_pdf_words(words: list[dict], *, digits_only: bool = False) -> str:
    ordered = sorted(words, key=lambda w: (round(float(w.get("top", 0)) / 3), float(w.get("x0", 0))))
    text = " ".join(normalize_text(w.get("text", "")) for w in ordered if normalize_text(w.get("text", "")))
    if digits_only:
        return re.sub(r"[^0-9*]", "", text)
    text = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])", "", text)
    return normalize_text(text)


def pdf_words_in_column(words: list[dict], left: float, right: float) -> list[dict]:
    selected = []
    for word in words:
        x0 = float(word.get("x0", 0))
        x1 = float(word.get("x1", x0))
        center = (x0 + x1) / 2
        if left <= center < right:
            selected.append(word)
    return selected


def extract_corporate_query_pdf(
    path: Path,
    passwords: list[str] | None = None,
    should_cancel=None,
) -> tuple[list[dict], str]:
    passwords = passwords or [""]
    account = infer_account_from_filename(path.name) or normalize_source_account(path.name)
    last_error = ""
    for password in passwords:
        txns: list[dict] = []
        try:
            with pdfplumber.open(path, password=password or None) as pdf:
                for page_no, page in enumerate(pdf.pages, start=1):
                    raise_if_cancelled(should_cancel)
                    words = page.extract_words(x_tolerance=1, y_tolerance=3, use_text_flow=False) or []
                    starts = [
                        word
                        for word in words
                        if float(word.get("x0", 999)) < 80
                        and re.fullmatch(r"20\d{2}-\d{2}-\d{2}", normalize_text(word.get("text", "")))
                    ]
                    starts.sort(key=lambda w: float(w.get("top", 0)))
                    for idx, start in enumerate(starts):
                        start_top = float(start.get("top", 0))
                        end_top = float(starts[idx + 1].get("top", page.height)) if idx + 1 < len(starts) else float(page.height)
                        block_top = max(0.0, start_top - 10.0)
                        block_bottom = max(block_top + 1.0, end_top - 10.0)
                        block = [
                            word
                            for word in words
                            if block_top <= float(word.get("top", 0)) < block_bottom
                            and "页码" not in normalize_text(word.get("text", ""))
                            and not re.fullmatch(r"\d+/\d+", normalize_text(word.get("text", "")))
                        ]
                        time_words = [
                            word
                            for word in pdf_words_in_column(block, 0, 90)
                            if re.fullmatch(r"\d{2}:\d{2}:\d{2}", normalize_text(word.get("text", "")))
                        ]
                        date_text = normalize_text(start.get("text", ""))
                        time_text = normalize_text(time_words[0].get("text", "")) if time_words else ""
                        debit = money_to_float(join_pdf_words(pdf_words_in_column(block, 90, 170)))
                        credit = money_to_float(join_pdf_words(pdf_words_in_column(block, 170, 250)))
                        balance = money_to_float(join_pdf_words(pdf_words_in_column(block, 250, 325)))
                        summary = join_pdf_words(pdf_words_in_column(block, 325, 392))
                        counterparty = clean_counterparty_text(join_pdf_words(pdf_words_in_column(block, 392, 462)))
                        counterparty_account = join_pdf_words(pdf_words_in_column(block, 462, 528), digits_only=True)
                        txn_type = join_pdf_words(pdf_words_in_column(block, 528, 590))
                        if debit is None and credit is None:
                            continue
                        summary_parts = [summary, txn_type]
                        if counterparty_account:
                            summary_parts.append(f"对方账号:{counterparty_account}")
                        if time_text:
                            summary_parts.append(time_text)
                        txn = make_txn(
                            date=parse_date(date_text),
                            account=account,
                            counterparty=counterparty or counterparty_account or txn_type or "未识别",
                            summary=" ".join(part for part in summary_parts if part),
                            income=credit,
                            expense=debit,
                            amount=None,
                            balance=balance,
                            source=path.name,
                            raw_key=f"{page_no}:{idx}:{date_text}:{time_text}:{debit}:{credit}:{balance}:{counterparty}:{counterparty_account}:{summary}:{txn_type}",
                            preserve_signed_columns=True,
                        )
                        if txn:
                            txns.append(txn)
            return dedupe_transactions(txns), password
        except Exception as exc:
            last_error = str(exc)
            if not re.search(r"password|encrypted|加密|密码", last_error, re.I):
                raise
    if len(passwords) <= 1 and not passwords[0]:
        raise PasswordRequiredError(path.name)
    raise PasswordRequiredError(path.name, f"{path.name} 密码不正确，或还需要输入正确密码。")


def extract_abc_account_detail_pdf(
    path: Path,
    passwords: list[str] | None = None,
    should_cancel=None,
) -> tuple[list[dict], str]:
    """农业银行账户明细/明细回单固定列 PDF。

    这类 PDF 的文字层会把同一行拆成多条视觉行；如果走通用文本规则，
    “对方账号”会被误识别成交易金额。这里按页面坐标列解析。
    """
    passwords = passwords or [""]
    last_error = ""
    for password in passwords:
        txns: list[dict] = []
        account = infer_account_from_filename(path.name) or normalize_source_account(path.name)
        try:
            with pdfplumber.open(path, password=password or None) as pdf:
                for page_no, page in enumerate(pdf.pages, start=1):
                    raise_if_cancelled(should_cancel)
                    page_text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
                    header_account = re.search(r"账号[:：]\s*([0-9A-Za-z* -]{8,})", page_text)
                    if header_account:
                        account = normalize_account(header_account.group(1), account)
                    words = page.extract_words(x_tolerance=1, y_tolerance=3, use_text_flow=False) or []
                    starts = [
                        word
                        for word in words
                        if float(word.get("x0", 999)) < 82
                        and re.fullmatch(r"20\d{2}-\d{2}-\d{2}", normalize_text(word.get("text", "")))
                    ]
                    starts.sort(key=lambda w: float(w.get("top", 0)))
                    for idx, start in enumerate(starts):
                        date_text = normalize_text(start.get("text", ""))
                        start_top = float(start.get("top", 0))
                        next_top = float(starts[idx + 1].get("top", page.height)) if idx + 1 < len(starts) else float(page.height)
                        block_top = max(72.0, start_top - 10.0)
                        block_bottom = max(block_top + 1.0, next_top - 8.0)
                        block = [
                            word
                            for word in words
                            if block_top <= float(word.get("top", 0)) < block_bottom
                            and not re.search(r"账户明细|账号[:：]|交易时间|收入金额|第\d+页", normalize_text(word.get("text", "")))
                        ]
                        time_words = [
                            word
                            for word in pdf_words_in_column(block, 0, 82)
                            if re.fullmatch(r"\d{2}:\d{2}:\d{2}", normalize_text(word.get("text", "")))
                        ]
                        time_text = normalize_text(time_words[0].get("text", "")) if time_words else ""
                        income = money_to_float(join_pdf_words(pdf_words_in_column(block, 82, 145)))
                        expense = money_to_float(join_pdf_words(pdf_words_in_column(block, 145, 205)))
                        balance = money_to_float(join_pdf_words(pdf_words_in_column(block, 205, 265)))
                        counterparty_account = join_pdf_words(pdf_words_in_column(block, 265, 330))
                        counterparty = clean_counterparty_text(join_pdf_words(pdf_words_in_column(block, 330, 392)))
                        purpose = join_pdf_words(pdf_words_in_column(block, 392, 455))
                        counterparty_bank = join_pdf_words(pdf_words_in_column(block, 455, 525))
                        summary = join_pdf_words(pdf_words_in_column(block, 525, 590))
                        if income is None and expense is None:
                            continue
                        summary_parts = [purpose, summary, counterparty_bank]
                        if counterparty_account:
                            summary_parts.append(f"对方账号:{counterparty_account}")
                        if time_text:
                            summary_parts.append(time_text)
                        txn = make_txn(
                            date=parse_date(date_text),
                            account=account,
                            counterparty=counterparty or counterparty_account or counterparty_bank or "未识别",
                            summary=" ".join(part for part in summary_parts if part),
                            income=income,
                            expense=expense,
                            amount=None,
                            balance=balance,
                            source=path.name,
                            raw_key=f"abc:{page_no}:{idx}:{date_text}:{time_text}:{income}:{expense}:{balance}:{counterparty_account}:{counterparty}:{purpose}:{summary}",
                            preserve_signed_columns=True,
                        )
                        if txn:
                            txns.append(txn)
            return dedupe_transactions(txns), password
        except Exception as exc:
            last_error = str(exc)
            if not re.search(r"password|encrypted|加密|密码", last_error, re.I):
                raise
    if len(passwords) <= 1 and not passwords[0]:
        raise PasswordRequiredError(path.name)
    raise PasswordRequiredError(path.name, f"{path.name} 密码不正确，或还需要输入正确密码。")


def old_extract_pdf_text_unused(path: Path) -> str:
    chunks = []
    try:
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
                chunks.append(f"【第 {i} 页】\n{text}")
                for table in page.extract_tables() or []:
                    for row in table:
                        chunks.append(" | ".join(normalize_text(cell) for cell in row or []))
    except Exception:
        return ""
    return "\n".join(chunks)


def ocr_file(path: Path, job_dir: Path, progress=None, passwords: list[str] | None = None, should_cancel=None) -> str:
    raise_if_cancelled(should_cancel)
    passwords = passwords or [""]
    for password in passwords:
        cache = CACHE / f"{cache_key(path, password)}.txt"
        if cache.exists():
            if progress:
                progress(72, "复用上次 OCR 结果")
            return cache.read_text(encoding="utf-8", errors="ignore")
    suffix = path.suffix.lower()
    if suffix in PDF_EXTS:
        if progress:
            progress(15, "渲染 PDF 页面")
        rendered = []
        used_password = ""
        errors = []
        for password in passwords:
            raise_if_cancelled(should_cancel)
            try:
                def render_progress(done, total, message):
                    raise_if_cancelled(should_cancel)
                    if progress:
                        progress(15 + int(10 * done / max(total, 1)), message)

                rendered = render_pdf(path, job_dir / "rendered", progress_cb=render_progress, password=password)
                if rendered:
                    used_password = password
                    break
            except Exception as exc:
                errors.append(str(exc))
                rendered = []
        if not rendered:
            raise PasswordRequiredError(path.name, f"{path.name} 需要输入正确的 PDF 密码，或该 PDF 无法渲染。")
        images = []
        for page_no, image in enumerate(rendered, start=1):
            raise_if_cancelled(should_cancel)
            if progress:
                progress(20 + int(20 * page_no / max(len(rendered), 1)), f"图像增强 {page_no}/{len(rendered)} 页")
            images.extend(prepare_and_split(image, job_dir / "split", page_no, include_wide_splits=False))
    else:
        raise_if_cancelled(should_cancel)
        if progress:
            progress(25, "图像增强")
        images = prepare_and_split(path, job_dir / "split", 1, include_wide_splits=False)

    def ocr_progress(done, total, _img):
        raise_if_cancelled(should_cancel)
        if progress:
            page_match = re.search(r"p(\d+)_", Path(str(_img)).stem)
            page_no = int(page_match.group(1)) if page_match else 1
            page_total = len(rendered) if suffix in PDF_EXTS else 1
            variant_no = ((done - 1) % 4) + 1
            progress(
                45 + int(40 * done / max(total, 1)),
                f"OCR 识别第 {page_no}/{page_total} 页（增强 {variant_no}/4）",
            )

    pages = ocr_images(
        images,
        progress_cb=ocr_progress if progress or should_cancel else None,
    )
    pages = select_best_ocr_pages(pages)
    raise_if_cancelled(should_cancel)
    text = "\n".join(f"【第 {page['page']} 页】\n{page['text']}" for page in pages)
    CACHE.mkdir(parents=True, exist_ok=True)
    cache_password = used_password if suffix in PDF_EXTS else ""
    (CACHE / f"{cache_key(path, cache_password)}.txt").write_text(text, encoding="utf-8")
    return text


def ocr_variant_key(image_path: str) -> str:
    stem = Path(image_path).stem
    stem = re.sub(r"_(?:nowm|colorclean|bw)$", "", stem)
    return stem


def ocr_text_score(text: str) -> int:
    lines = [normalize_text(line) for line in text.splitlines() if normalize_text(line)]
    transaction_rows = 0
    for line in lines:
        if parse_date(line) and extract_signed_amounts_from_line(line):
            transaction_rows += 1
    table_headers = sum(1 for line in lines if is_table_header(line))
    useful_dates = sum(1 for line in lines if parse_date(line))
    return transaction_rows * 1000 + table_headers * 200 + useful_dates * 10 + min(len(text), 500)


def select_best_ocr_pages(pages: list[dict]) -> list[dict]:
    groups: dict[str, dict] = {}
    order: list[str] = []
    for page in pages:
        key = ocr_variant_key(page.get("image", ""))
        if key not in groups:
            groups[key] = page
            order.append(key)
            continue
        if ocr_text_score(page.get("text", "")) > ocr_text_score(groups[key].get("text", "")):
            groups[key] = page
    return [groups[key] for key in order]


def normalize_money_fragments(line: str) -> str:
    text = normalize_text(line)
    text = re.sub(r"([+-]?\d{1,3})[,.，]\s*(\d{3})[,.，]\s*(\d{2})(?!\d)", r"\1,\2.\3", text)
    text = re.sub(r"([+-]?\d{1,3})[,.，]\s*(\d{3})(\d{2})(?!\d)", r"\1,\2.\3", text)
    text = re.sub(r"([+-]?\d{1,3})[,，]\s+(\d{3})", r"\1,\2", text)
    text = re.sub(r"([+-]?\d+)[.]\s+(\d{2})(?!\d)", r"\1.\2", text)
    return text


def extract_signed_amounts_from_line(line: str) -> list[float]:
    text = normalize_money_fragments(line)
    values = []
    pattern = r"(?<!\d)[+-]\s*(?:\d{1,3}(?:[,，]\d{3})+|\d+)(?:\.\d{1,2})?(?!\d)"
    for match in re.finditer(pattern, text):
        value = money_to_float(match.group(0).replace(" ", ""))
        if value is not None:
            values.append(value)
    return values


def extract_amounts_from_line(line: str) -> list[float]:
    line = normalize_money_fragments(line)
    values = []
    for match in re.finditer(r"(?<!\d)[+-]?(?:\d{1,3}(?:[,，]\d{3})+|\d+)(?:\.\d{1,2})?(?!\d)", line):
        value = money_to_float(match.group(0))
        if value is not None:
            values.append(value)
    return values


def clean_counterparty_text(text: str) -> str:
    text = normalize_text(text)
    text = re.sub(r"^(?:[A-Fa-f0-9]\s*){1,8}(?=[\u4e00-\u9fff])", "", text)
    text = re.sub(r"(?<=[\u4e00-\u9fff])[A-Fa-f0-9](?=[\u4e00-\u9fff])", "", text)
    text = re.sub(r'(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])', '', text)
    text = re.sub(r"\b[A-Z]\d{6,}\b.*$", "", text)
    text = re.sub(r"\b(?:电子商务|超级网银|掌上银行|柜面|网上银行|自助终端|其他)\b.*$", "", text)
    text = re.sub(r"\b[0-9A-Z]{10,}\b.*$", "", text)
    text = normalize_text(text)
    return text or "未识别"


def infer_counterparty_after_amount(line: str) -> str:
    text = normalize_money_fragments(line)
    signed = re.search(r"(?<!\d)[+-]\s*(?:\d{1,3}(?:[,，]\d{3})+|\d+)(?:\.\d{1,2})?(?!\d)", text)
    if not signed:
        return ""
    tail = text[signed.end() :]
    tail = re.sub(r"^\s*(?:\d{1,3}(?:[,，]\d{3})+|\d+)(?:\.\d{1,2})?\s+", " ", tail)
    parts = [normalize_text(p) for p in re.split(r"\s+|\|", tail) if normalize_text(p)]
    skip = re.compile(r"^\d|^\*+$|^[0-9*]{4,}$|手机银行|网上银行|快捷支付|电子商务|其他|掌上银行|超级网银|自动柜员机")
    candidates = [p for p in parts if len(p) >= 2 and not skip.search(p) and money_to_float(p) is None]
    return clean_counterparty_text(candidates[0]) if candidates else ""


def is_non_transaction_line(line: str) -> bool:
    if is_summary_text(line):
        return True
    return bool(
        re.search(
            r"编号[:：]|兹证明|居民身份证|交易明细对应时间段|具体交易明细|交易单号|交易日期\s*交易时间|"
            r"户名[:：]|币种[:：]|起止日期[:：]|时间段[:：]|开立日期[:：]|电子流水号|第\d+页|共\d+页|"
            r"该交易明细|中国农业银行账户活期交易明细清单|中国工商银行借记账户|"
            r"中国银行交易流水明细清单|交易区间[:：]|客户姓名[:：]|借记卡号[:：]|按收支筛选[:：]|打印时间[:：]|"
            r"业务专用章|记账日期|对方卡号[/／]账号|"
            r"卡号\d*$|^账号$|^储种$|^序号$|^币种$|^摘要$|^地区$|^收[/／]?支|^余额$|"
            r"^\d{12,}$|^\d+-\d{10,}$|^[A-Z0-9]{8,}$",
            line,
        )
    )


def parse_alipay_table_line(line: str, account: str, source: str) -> dict | None:
    if "|" not in line:
        return None
    cells = [normalize_text(cell) for cell in line.split("|")]
    while cells and not cells[0]:
        cells = cells[1:]
    while cells and not cells[-1]:
        cells = cells[:-1]
    if len(cells) < 8 or cells[0] in {"收/支", "交易类型：全部", "交易时间段："}:
        return None
    direction_text = normalize_text(cells[0]).replace(" ", "")
    if "不计" in direction_text:
        return None
    if "收入" not in direction_text and "支出" not in direction_text:
        return None
    date_index = next((idx for idx in range(len(cells) - 1, -1, -1) if parse_date(cells[idx])), -1)
    date = parse_date(cells[date_index]) if date_index >= 0 else None
    amount = None
    if date_index > 0:
        for cell in reversed(cells[:date_index]):
            if re.fullmatch(r"[+-]?(?:\d{1,3}(?:,\d{3})+|\d+)\.\d{1,2}", cell.replace(" ", "")):
                amount = money_to_float(cell)
                break
    if amount is None or not date:
        return None
    income = abs(amount) if "收入" in direction_text else None
    expense = abs(amount) if "支出" in direction_text else None
    return make_txn(
        date=date,
        account=account,
        counterparty=cells[1] or "未识别",
        summary=" ".join(cell for cell in [cells[2], cells[3]] if cell),
        income=income,
        expense=expense,
        amount=amount if income else -abs(amount),
        balance=None,
        source=source,
        raw_key=line,
    )


def parse_wechat_table_line(line: str, source: str, statement_account: str = "") -> dict | None:
    if "|" not in line:
        return None
    cells = [normalize_text(cell) for cell in line.split("|")]
    if is_summary_row(cells):
        return None
    if len(cells) < 7 or cells[0] in {"交易单号", "具体交易明细", "交易明细对应时间段"}:
        return None
    date = parse_date(cells[1])
    amount = money_to_float(cells[5] if len(cells) > 5 else "")
    if not date or amount is None:
        return None
    direction_text = cells[3] if len(cells) > 3 else ""
    direction_clean = normalize_text(direction_text).replace(" ", "")
    is_internal = "其他" in direction_clean
    direction = detect_direction(direction_text, amount)
    income = abs(amount) if direction == "收入" or is_internal else None
    expense = abs(amount) if direction == "支出" and not is_internal else None
    if not income and not expense:
        return None
    method = cells[4] if len(cells) > 4 else ""
    counterparty = cells[6] if len(cells) > 6 else "未识别"
    summary = " ".join(cell for cell in ["微信内部流水" if is_internal else "", cells[2] if len(cells) > 2 else "", direction_text, method] if cell)
    account = statement_account or normalize_source_account(source)
    txn = make_txn(
        date=date,
        account=account,
        counterparty=counterparty,
        summary=summary,
        income=income,
        expense=expense,
        amount=amount if income else -abs(amount),
        balance=None,
        source=source,
        raw_key=line,
    )
    if txn and is_internal:
        txn["category"] = "微信内部流水"
        txn["default_excluded"] = True
        txn["exclude_reason"] = "微信内部流水"
        txn["direction"] = "其他"
        txn["other_amount"] = round(abs(float(amount or 0)), 2)
        txn["income"] = 0.0
        txn["expense"] = 0.0
        txn["amount"] = 0.0
    return txn


def raw_pipe_cells(line: str) -> list[str]:
    return [normalize_text(cell) for cell in line.split("|")]


def noisy_signed_money(value: str) -> float | None:
    text = normalize_money_fragments(value)
    signed = []
    for match in re.finditer(r"[+-]\s*(?:\d{1,3}(?:[,，]\d{3})+|\d+)(?:\.\d{1,3})?", text):
        amount = money_to_float(match.group(0).replace(" ", ""))
        if amount is not None:
            signed.append(amount)
    if not signed:
        signed = extract_signed_amounts_from_line(text)
    if signed:
        return signed[-1]
    candidates = re.findall(r"(?:\d{1,3}(?:[,，]\d{3})+|\d+)\.\d{1,3}", text)
    return money_to_float(candidates[-1]) if candidates else None


def noisy_balance_money(value: str) -> float | None:
    text = normalize_money_fragments(value)
    candidates = re.findall(r"(?:\d{1,3}(?:[,，]\d{3})+|\d+)\.\d{1,3}", text)
    if not candidates:
        return None
    amount = money_to_float(candidates[-1])
    return round(float(amount), 2) if amount is not None else None


def noisy_datetime(value: str) -> str:
    # 工行防伪水印会把“行/银”等字插进年份（如“行 2银025”）。
    # 日期列只移除汉字扰码，保留原有数字和分隔符。
    cleaned = re.sub(r"[\u4e00-\u9fff]", "", normalize_text(value))
    match = re.search(r"20\d{2}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}", cleaned)
    return match.group(0) if match else ""


def is_grcb_broken_prefix_line(line: str) -> bool:
    cells = raw_pipe_cells(line)
    return (
        len(cells) >= 9
        and not cells[0]
        and cells[1]
        and not cells[2]
        and not cells[3]
        and re.fullmatch(r"\d+\.", cells[4] or "") is not None
        and re.fullmatch(r"\d+\.", cells[5] or "") is not None
    )


def merge_grcb_broken_line(prefix_line: str, line: str) -> str | None:
    prefix = raw_pipe_cells(prefix_line)
    cells = raw_pipe_cells(line)
    if not (
        len(prefix) >= 9
        and len(cells) >= 13
        and re.fullmatch(r"\d{1,8}", cells[0] or "")
        and re.match(r"20\d{2}[-/.]\d{1,2}[-/.]\d{1,2}", cells[2] or "")
        and not cells[3]
        and re.fullmatch(r"\d{1,2}", cells[4] or "")
        and re.fullmatch(r"\d{1,2}", cells[5] or "")
    ):
        return None
    merged = [
        cells[0],
        " ".join(cell for cell in [prefix[1], cells[1]] if cell),
        cells[2],
        cells[3],
        f"{prefix[4]} {cells[4]}",
        f"{prefix[5]} {cells[5]}",
        " ".join(cell for cell in [prefix[6], cells[6]] if cell),
        " ".join(cell for cell in [prefix[7], cells[7]] if cell),
        " ".join(cell for cell in [prefix[8], cells[8]] if cell),
        cells[9],
        cells[10],
        cells[11],
        cells[12],
    ]
    return " | ".join(merged)


def parse_bank_pipe_table_line(line: str, account: str, source: str) -> dict | None:
    if "|" not in line:
        return None
    cells = [normalize_text(cell) for cell in line.split("|")]
    while cells and not cells[0]:
        cells = cells[1:]
    while cells and not cells[-1]:
        cells = cells[:-1]
    if is_summary_row(cells) or len(cells) < 5:
        return None
    header_text = "".join(cells[:5])
    if re.search(r"交易时\s*间|交易时间|交易日期|会计日期|交易金额|流水号收入|借方发生额|贷方发生额", header_text):
        return None

    # 邮储银行历史明细：交易时间 | 子账号 | 储种 | 币种 | 钞汇 | 交易金额 | 交易余额 | 对方户名 | 对方账号 | 摘要 | 交易渠道 | 外部系统流水。
    # 这类表格外形接近工行个人借记卡历史明细；必须先于工行规则判断，
    # 否则第 9 列“对方账号”会被工行分支误当作金额，导致整行被丢弃。
    if (
        len(cells) >= 10
        and re.fullmatch(r"20\d{2}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}", cells[0])
        and re.fullmatch(r"\d{4}", cells[1])
        and "活期" in cells[2]
        and "人民币" in cells[3]
        and money_to_float(cells[5]) is not None
        and money_to_float(cells[6]) is not None
    ):
        amount = money_to_float(cells[5])
        balance = money_to_float(cells[6])
        return make_txn(
            date=parse_date(cells[0]),
            account=account,
            counterparty=cells[7] or cells[8] or "未识别",
            summary=" ".join(cell for cell in [cells[9] if len(cells) > 9 else "", cells[10] if len(cells) > 10 else ""] if cell),
            income=amount if amount and amount > 0 else None,
            expense=abs(amount) if amount and amount < 0 else None,
            amount=amount,
            balance=balance,
            source=source,
            raw_key=line,
        )

    # 工商银行个人借记卡历史明细：日期时间 | 账号 | 储种 | 序号 | 币种 | 钞汇 | 摘要 | 地区 | 收支金额 | 余额 | 渠道。
    # 工行电子版带防伪扰码，金额列可能混入单个字符或数字，必须锁定第9/10列，不能走通用数字推断。
    icbc_time = noisy_datetime(cells[0]) if cells else ""
    if len(cells) >= 11 and icbc_time and ("活期" in cells[2] or re.search(r"\d{15,}", cells[1])):
        amount = noisy_signed_money(cells[8])
        balance = noisy_balance_money(cells[9])
        if amount is None:
            return None
        counterparty = clean_counterparty_text(cells[10] if len(cells) > 10 else "")
        counterparty_account = normalize_text(cells[11] if len(cells) > 11 else "")
        channel = normalize_text(cells[12] if len(cells) > 12 else "")
        return make_txn(
            date=parse_date(icbc_time),
            account=account,
            counterparty=counterparty or counterparty_account or "工商银行",
            summary=" ".join(cell for cell in [cells[6], channel, f"对方账号:{counterparty_account}" if counterparty_account else ""] if cell),
            income=amount if amount > 0 else None,
            expense=abs(amount) if amount < 0 else None,
            amount=amount,
            balance=balance,
            source=source,
            raw_key=line,
        )

    # 中国银行个人流水：日期 | 时间 | 币种 | 带符号金额 | 余额 | 交易名称 | 渠道 | 网点 | 附言 | 对方户名 | 对方账号 | 对方行。
    if (
        len(cells) >= 10
        and re.fullmatch(r"20\d{2}-\d{2}-\d{2}", cells[0])
        and re.fullmatch(r"\d{2}:\d{2}:\d{2}", cells[1])
        and "人民币" in cells[2]
    ):
        amount = money_to_float(cells[3])
        balance = money_to_float(cells[4])
        if amount is None or balance is None:
            return None
        counterparty = (cells[9] if len(cells) > 9 else "") or (cells[8] if len(cells) > 8 else "") or "未识别"
        return make_txn(
            date=parse_date(cells[0]),
            account=account,
            counterparty=clean_counterparty_text(counterparty),
            summary=" ".join(cell for cell in [cells[5], cells[6], cells[8] if len(cells) > 8 else ""] if cell),
            income=amount if amount > 0 else None,
            expense=abs(amount) if amount < 0 else None,
            amount=amount,
            balance=balance,
            source=source,
            raw_key=line,
        )

    # 平安银行个人账户流水：序号 | 交易日期 | 带符号金额 | 余额 | 交易地点 | 摘要 | 备注
    pingan_date = re.sub(r"(?i)[ABP\s]", "", cells[1]) if len(cells) > 1 else ""
    if len(cells) >= 6 and re.search(r"\d", cells[0]) and re.fullmatch(r"20\d{2}-\d{2}-\d{2}", pingan_date):
        amount_text = re.sub(r"(?i)[ABP\s]", "", cells[2])
        balance_text = re.sub(r"(?i)[ABP\s]", "", cells[3])
        amount = money_to_float(amount_text)
        balance = money_to_float(balance_text)
        if amount is not None and balance is not None:
            return make_txn(
                date=parse_date(pingan_date),
                account=account,
                counterparty=cells[4] or "平安银行",
                summary=" ".join(cell for cell in cells[5:] if cell),
                income=amount if amount > 0 else None,
                expense=abs(amount) if amount < 0 else None,
                amount=amount,
                balance=balance,
                source=source,
                raw_key=line,
            )

    # 农商银行账户对账单：流水号 | 交易日期 | 时间 | 摘要 | 支票号 | 支出 | 收入 | 余额 | 机构号
    if (
        len(cells) >= 8
        and re.search(r"\d", cells[0])
        and re.fullmatch(r"20\d{6}", cells[1])
        and re.fullmatch(r"\d{2}:\d{2}:\d{2}", cells[2])
    ):
        expense = money_to_float(cells[5])
        income = money_to_float(cells[6])
        balance = money_to_float(cells[7])
        if (income and income > 0) or (expense and expense > 0):
            return make_txn(
                date=parse_date(cells[1]),
                account=account,
                counterparty=cells[8] if len(cells) > 8 and cells[8] else "未识别",
                summary=cells[3],
                income=income,
                expense=expense,
                amount=None,
                balance=balance,
                source=source,
                raw_key=line,
            )

    # 工商银行账户明细清单：交易时间 | 转入金额 | 转出金额 | 余额 | 对方账号 | 对方单位 | 对方行名 | 摘要
    if (
        len(cells) >= 7
        and re.fullmatch(r"20\d{2}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}", cells[0])
        and money_to_float(cells[3]) is not None
        and (not cells[1] or money_to_float(cells[1]) is not None)
        and (not cells[2] or money_to_float(cells[2]) is not None)
    ):
        income = money_to_float(cells[1])
        expense = money_to_float(cells[2])
        balance = money_to_float(cells[3])
        if (income and income > 0) or (expense and expense > 0):
            return make_txn(
                date=parse_date(cells[0]),
                account=account,
                counterparty=cells[5] or cells[4] or cells[6] or "未识别",
                summary=" ".join(cell for cell in [cells[7] if len(cells) > 7 else "", cells[6]] if cell),
                income=income,
                expense=expense,
                amount=None,
                balance=balance,
                source=source,
                raw_key=line,
            )

    # 邮储银行：交易时间 | 子账号 | 储种 | 币种 | 钞汇 | 交易金额 | 交易余额 | 对方户名 | 对方账号 | 摘要 | ...
    if (
        len(cells) >= 10
        and re.fullmatch(r"20\d{2}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}", cells[0])
        and re.fullmatch(r"\d{4}", cells[1])
        and "活期" in cells[2]
    ):
        amount = money_to_float(cells[5])
        balance = money_to_float(cells[6])
        if amount is None:
            return None
        return make_txn(
            date=parse_date(cells[0]),
            account=account,
            counterparty=cells[7] or cells[8] or "未识别",
            summary=" ".join(cell for cell in cells[9:11] if cell),
            income=amount if amount > 0 else None,
            expense=abs(amount) if amount < 0 else None,
            amount=amount,
            balance=balance,
            source=source,
            raw_key=line,
        )

    # 浦发银行个人流水：交易日期 | 时间 | 本方账号 | 交易名称 | 交易金额 | 余额 | 对手姓名 | 对手账号 | 摘要
    if len(cells) >= 8 and re.fullmatch(r"20\d{6}", cells[0]) and re.fullmatch(r"\d{6}", cells[1]):
        amount = money_to_float(cells[4])
        balance = money_to_float(cells[5])
        if amount is None:
            return None
        return make_txn(
            date=parse_date(cells[0]),
            account=normalize_account(cells[2], account),
            counterparty=cells[6] or cells[7] or "未识别",
            summary=" ".join(cell for cell in [cells[3], cells[8] if len(cells) > 8 else ""] if cell),
            income=amount if amount > 0 else None,
            expense=abs(amount) if amount < 0 else None,
            amount=amount,
            balance=balance,
            source=source,
            raw_key=line,
        )

    # 招商银行：交易时间 | 交易金额（带正负号） | 余额 | 交易类型 | 交易备注
    if len(cells) == 5 and re.fullmatch(r"20\d{2}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}", cells[0]):
        amount = money_to_float(cells[1])
        balance = money_to_float(cells[2])
        if amount is None or balance is None:
            return None
        txn_type = cells[3]
        remark = " ".join(cell for cell in cells[4:] if cell)
        counterparty_text = re.sub(
            r"^(?:转账|往来款|货款|订金|生活费|跨行转出|跨行网银贷记业务)\s*",
            "",
            remark,
        ).strip()
        if "-" in counterparty_text:
            parts = [part.strip() for part in counterparty_text.split("-") if part.strip()]
            organization_parts = [part for part in parts if re.search(r"公司|银行|合作社|商行|租赁|保险", part)]
            generic_parts = {'代付', '还款', '微信零钱提现', '支付宝余额提现', '微信支付', '扫二维码付款'}
            named_parts = [part for part in parts if re.fullmatch(r"[\u4e00-\u9fff·]{2,12}", part) and part not in generic_parts]
            if organization_parts:
                counterparty_text = organization_parts[-1]
            elif named_parts:
                counterparty_text = named_parts[-1]
            elif parts:
                counterparty_text = parts[-1]
        counterparty = clean_counterparty_text(counterparty_text or txn_type)
        return make_txn(
            date=parse_date(cells[0]),
            account=account,
            counterparty=counterparty,
            summary=" ".join(part for part in [txn_type, remark] if part),
            income=amount if amount > 0 else None,
            expense=abs(amount) if amount < 0 else None,
            amount=amount,
            balance=balance,
            source=source,
            raw_key=line,
        )

    # 广州农商银行: 序号 | 电子回单号 | 交易日期 | 支出 | 存入 | 账户余额 | 对方账号 | 对方户名 | 对方行名称 | 摘要 | 用途/其他摘要 | 自然日期 | 自然时间
    if len(cells) >= 13 and re.fullmatch(r"\d{1,8}", cells[0]) and re.match(r"20\d{2}[-/.]\d{1,2}[-/.]\d{1,2}", cells[2]):
        expense = money_to_float(cells[3])
        income = money_to_float(cells[4])
        balance = money_to_float(cells[5])
        if not income and not expense:
            return None
        return make_txn(
            date=parse_date(cells[2]),
            account=account,
            counterparty=cells[7] or cells[6] or "未识别",
            summary=" ".join(cell for cell in [cells[9], cells[10]] if cell),
            income=income,
            expense=expense,
            amount=None,
            balance=balance,
            source=source,
            raw_key=line,
        )

    # 中国银行: 序号 | 记账日 | 起息日 | 交易类型 | 凭证 | 凭证号码/业务编号/用途/摘要 | 借方发生额 | 贷方发生额 | 余额 | 机构/柜员/流水 | 备注
    if len(cells) >= 10 and re.fullmatch(r"\d{1,8}", cells[0]) and re.fullmatch(r"\d{6}", cells[1]):
        expense = money_to_float(cells[6])
        income = money_to_float(cells[7])
        balance = money_to_float(cells[8])
        if not income and not expense:
            return None
        return make_txn(
            date=parse_short_bank_date(cells[1]),
            account=account,
            counterparty=(cells[10] if len(cells) > 10 else "") or cells[9] or "未识别",
            summary=" ".join(cell for cell in [cells[3], cells[5]] if cell),
            income=income,
            expense=expense,
            amount=None,
            balance=balance,
            source=source,
            raw_key=line,
        )

    # 交通银行: 序号 | 会计日期 | 交易日期 | 交易名称 | 凭证种类 | 凭证号码 | 借方发生额 | 贷方发生额 | 余额 | 卡号 | 交易地点 | 对方账号 | 对方户名 | 对方行名 | 摘要 | 流水号
    if len(cells) >= 16 and re.fullmatch(r"\d{1,8}", cells[0]) and re.fullmatch(r"20\d{6}", cells[1]) and re.fullmatch(r"20\d{6}", cells[2]):
        expense = money_to_float(cells[6])
        income = money_to_float(cells[7])
        balance = money_to_float(cells[8])
        if not income and not expense:
            return None
        return make_txn(
            date=parse_date(cells[2]),
            account=account,
            counterparty=cells[12] or cells[11] or cells[13] or "未识别",
            summary=" ".join(cell for cell in [cells[3], cells[14]] if cell),
            income=income,
            expense=expense,
            amount=None,
            balance=balance,
            source=source,
            raw_key=line,
        )

    # 农商行账户明细: 流水号 | 交易日期 | 收入 | 支出 | 账户余额 | 对方账号 | 对方户名 | 对方行名 | ... | 交易类型 | 支票号 | 附言
    if len(cells) >= 10 and re.fullmatch(r"[\d ]{4,}", cells[0]) and re.match(r"20\d{2}[-/.]\d{1,2}[-/.]\d{1,2}", cells[1]):
        income = money_to_float(cells[2])
        expense = money_to_float(cells[3])
        balance = money_to_float(cells[4])
        if not income and not expense:
            return None
        return make_txn(
            date=parse_date(cells[1]),
            account=account,
            counterparty=cells[6] or cells[5] or cells[7] or "未识别",
            summary=" ".join(cell for cell in [cells[9] if len(cells) > 9 else "", cells[11] if len(cells) > 11 else ""] if cell),
            income=income,
            expense=expense,
            amount=None,
            balance=balance,
            source=source,
            raw_key=line,
        )

    # 建设银行账户明细: 账号 | 交易时间 | 借方发生额 | 贷方发生额 | 余额 | 币种 | 对方户名 | 对方账号 | 对方开户机构 | 记账日期 | 摘要 | 备注
    if len(cells) >= 11 and re.fullmatch(r"[\d ]{8,}", cells[0]) and re.match(r"20\d{6}(?:\s+\d{1,2}:\d{2}:\d{2})?", cells[1]):
        expense = money_to_float(cells[2])
        income = money_to_float(cells[3])
        balance = money_to_float(cells[4])
        if not income and not expense:
            return None
        return make_txn(
            date=parse_date(cells[1]),
            account=normalize_account(cells[0], account),
            counterparty=cells[6] or cells[7] or cells[8] or "未识别",
            summary=" ".join(cell for cell in [cells[10] if len(cells) > 10 else "", cells[11] if len(cells) > 11 else ""] if cell),
            income=income,
            expense=expense,
            amount=None,
            balance=balance,
            source=source,
            raw_key=line,
        )

    # 建设银行: 序号 | 摘要 | 交易日期 | 交易金额 | 账户余额 | 交易地点/附言 | 对方账号与户名
    if len(cells) >= 5 and re.fullmatch(r"\d{1,8}", cells[0]) and re.fullmatch(r"20\d{6}", cells[2]):
        amount = money_to_float(cells[3])
        balance = money_to_float(cells[4])
        if amount is None:
            return None
        location = cells[5] if len(cells) > 5 else ""
        return make_txn(
            date=parse_date(cells[2]),
            account=account,
            counterparty=(cells[6] if len(cells) > 6 else "") or location or "未识别",
            summary=" ".join(cell for cell in [cells[1], location] if cell),
            income=amount if amount > 0 else None,
            expense=abs(amount) if amount < 0 else None,
            amount=amount,
            balance=balance,
            source=source,
            raw_key=line,
        )

    # 工商银行: 凭证号 | 对方账号 | 交易时间 | 借贷标志 | 对方单位 | 用途 | 附言 | 转出金额 | 转入金额 | 入账日期 | 余额 | 对方行名
    if len(cells) >= 12 and re.match(r"\d{6,}", cells[0]) and re.match(r"20\d{2}[-/.]\d{1,2}[-/.]\d{1,2}", cells[2]):
        direction = detect_direction(cells[3])
        expense = money_to_float(cells[7])
        income = money_to_float(cells[8])
        if direction == "支出":
            income = None
        elif direction == "收入":
            expense = None
        balance = money_to_float(cells[10])
        counterparty = cells[4] or cells[11] or "未识别"
        return make_txn(
            date=parse_date(cells[2]),
            account=account,
            counterparty=counterparty,
            summary=" ".join(cell for cell in [cells[5], cells[6]] if cell),
            income=income,
            expense=expense,
            amount=None,
            balance=balance,
            source=source,
            raw_key=line,
        )

    # 工商银行企业账户明细: 凭证号 | 对方行名 | 对方账号 | 对方单位 | 交易时间 | 转入金额 | 转出金额 | 用途 | 摘要 | 附言
    if len(cells) >= 9 and re.fullmatch(r"\d{6,}", cells[0]) and re.match(r"20\d{2}[-/.]\d{1,2}[-/.]\d{1,2}", cells[4]):
        income = money_to_float(cells[5])
        expense = money_to_float(cells[6])
        if not income and not expense:
            return None
        return make_txn(
            date=parse_date(cells[4]),
            account=account,
            counterparty=cells[3] or cells[2] or cells[1] or "未识别",
            summary=" ".join(cell for cell in [cells[7] if len(cells) > 7 else "", cells[8] if len(cells) > 8 else "", cells[9] if len(cells) > 9 else ""] if cell),
            income=income,
            expense=expense,
            amount=None,
            balance=None,
            source=source,
            raw_key=line,
        )

    # 广发银行: 流水号 | 交易时间 | 收入 | 支出 | 余额 | 币种 | 对方账号 | 对方户名 | ... | 摘要 | 备注 | 附言
    if len(cells) >= 13 and re.match(r"\d{6,}", cells[0]) and re.match(r"20\d{2}[-/.]\d{1,2}[-/.]\d{1,2}", cells[1]):
        income = money_to_float(cells[2])
        expense = money_to_float(cells[3])
        balance = money_to_float(cells[4])
        if not income and not expense:
            return None
        return make_txn(
            date=parse_date(cells[1]),
            account=account,
            counterparty=cells[7] or "未识别",
            summary=" ".join(cell for cell in [cells[12] if len(cells) > 12 else "", cells[13] if len(cells) > 13 else "", cells[14] if len(cells) > 14 else ""] if cell),
            income=income,
            expense=expense,
            amount=None,
            balance=balance,
            source=source,
            raw_key=line,
            preserve_signed_columns=True,
        )

    # 兴业银行: 交易时间 | 摘要 | 凭证代号 | 支出 | 收入 | 收支方向 | 账户余额 | 对方账号 | 对方户名
    if (
        len(cells) >= 7
        and re.match(r"20\d{2}[-/.]\d{1,2}[-/.]\d{1,2}", cells[0])
        and len(cells) > 5
        and re.search(r"收入|支出|借|贷", cells[5])
    ):
        expense = money_to_float(cells[3])
        income = money_to_float(cells[4])
        balance = money_to_float(cells[6])
        if (income and income > 0) or (expense and expense > 0):
            return make_txn(
                date=parse_date(cells[0]),
                account=account,
                counterparty=(cells[8] if len(cells) > 8 else "") or (cells[7] if len(cells) > 7 else "") or "未识别",
                summary=cells[1],
                income=income,
                expense=expense,
                amount=None,
                balance=balance,
                source=source,
                raw_key=line,
            )

    # 邮储银行: 交易时间 | 子账号 | 储种 | 币种 | 钞汇 | 交易金额 | 交易余额 | 对方户名 | 对方账号 | 摘要 | 交易渠道 | 外部系统流水
    if len(cells) >= 10 and re.match(r"20\d{2}[-/.]\d{1,2}[-/.]\d{1,2}", cells[0]):
        amount = money_to_float(cells[5])
        balance = money_to_float(cells[6])
        if amount is None:
            return None
        return make_txn(
            date=parse_date(cells[0]),
            account=account,
            counterparty=cells[7] or "未识别",
            summary=" ".join(cell for cell in [cells[9], cells[10] if len(cells) > 10 else ""] if cell),
            income=amount if amount > 0 else None,
            expense=abs(amount) if amount < 0 else None,
            amount=amount,
            balance=balance,
            source=source,
            raw_key=line,
        )
    return None


def parse_abc_line(line: str, account: str, source: str) -> dict | None:
    match = re.match(
        r"^(?P<date>20\d{6})\s+(?:(?P<time>\d{6})\s+)?(?P<summary>.+?)\s+(?P<amount>[+-]?\d+(?:\.\d{1,2})?)\s+(?P<balance>[+-]?\d+(?:\.\d{1,2})?)\s+(?P<rest>.+)$",
        line,
    )
    if not match:
        return None
    date = parse_date(match.group("date"))
    amount = money_to_float(match.group("amount"))
    balance = money_to_float(match.group("balance"))
    if not date or amount is None:
        return None
    income = amount if amount > 0 else None
    expense = abs(amount) if amount < 0 else None
    rest = clean_counterparty_text(match.group("rest"))
    return make_txn(
        date=date,
        account=account,
        counterparty=rest,
        summary=match.group("summary"),
        income=income,
        expense=expense,
        amount=amount,
        balance=balance,
        source=source,
        raw_key=line,
    )


def parse_shunde_rcb_personal_text(text: str, account: str, source: str) -> list[dict]:
    lines = [normalize_text(line) for line in text.splitlines()]
    starts = [idx for idx, line in enumerate(lines) if re.match(r"^20\d{2}-\d{2}-\d{2}(?:\s|$)", line)]
    txns = []
    skip_names = {"转账", "核心渠道", "手机银行", "财付通", "付款", "协议支付", "贷款归还", "跨行网银贷记业务"}
    for position, start in enumerate(starts):
        end = starts[position + 1] if position + 1 < len(starts) else min(len(lines), start + 5)
        chunk_lines = [line for line in lines[start:end] if line]
        chunk = " ".join(chunk_lines)
        signed_match = re.search(r"(?<!\d)[+-]\s*(?:\d{1,3}(?:[,，]\d{3})+|\d+)(?:\.\d{1,2})?(?!\d)", normalize_money_fragments(chunk))
        if not signed_match:
            continue
        amount = money_to_float(signed_match.group(0).replace(" ", ""))
        if amount is None:
            continue
        tail = normalize_money_fragments(chunk)[signed_match.end():]
        balance_match = re.search(r"(?:\d{1,3}(?:[,，]\d{3})+|\d+)\.\d{1,2}", tail)
        balance = money_to_float(balance_match.group(0)) if balance_match else None
        party_area = tail[:balance_match.start()] if balance_match else tail
        party_area = re.sub(r"\b\d[\d ]{5,}\b|\*+", " ", party_area)
        names = [normalize_text(x) for x in re.findall(r"[\u4e00-\u9fff·]{2,30}", party_area)]
        names = [x for x in names if x not in skip_names and not re.search(r"银行股份有限公司|农村商业银行|中国银行总行", x)]
        counterparty = names[0] if names else "未识别"
        amount_line = next((line for line in chunk_lines if re.search(r"(?<!\d)[+-]\s*\d", line)), chunk_lines[0])
        summary = normalize_text(re.split(r"(?<!\d)[+-]\s*\d", amount_line, maxsplit=1)[0])
        summary = summary.split()[-1] if summary else "转账"
        txn = make_txn(
            date=parse_date(chunk_lines[0]),
            account=account,
            counterparty=counterparty,
            summary=summary,
            income=amount if amount > 0 else None,
            expense=abs(amount) if amount < 0 else None,
            amount=amount,
            balance=balance,
            source=source,
            raw_key=chunk,
        )
        if txn:
            txns.append(txn)
    return txns


def repair_running_balances(txns: list[dict], tolerance: float = 5.0) -> list[dict]:
    previous = None
    for txn in txns:
        income = float(txn.get("income") or 0)
        expense = float(txn.get("expense") or 0)
        balance = txn.get("balance")
        magnitude = round(income + expense, 2)
        if previous is not None and balance is not None and magnitude:
            delta = round(float(balance) - previous, 2)
            # 防伪字符偶尔落在金额正负号之间。余额差与金额绝对值吻合时，
            # 以连续余额反推真实收支方向和被扰码覆盖的尾数。
            if abs(abs(delta) - magnitude) <= tolerance:
                corrected = round(abs(delta), 2)
                if delta < 0:
                    txn["income"] = 0.0
                    txn["expense"] = corrected
                    txn["amount"] = -corrected
                    txn["direction"] = "支出"
                    income, expense = 0.0, corrected
                elif delta > 0:
                    txn["income"] = corrected
                    txn["expense"] = 0.0
                    txn["amount"] = corrected
                    txn["direction"] = "收入"
                    income, expense = corrected, 0.0
        expected = round(previous + income - expense, 2) if previous is not None else None
        if expected is not None and balance is not None and abs(float(balance) - expected) > tolerance:
            noisy_balance = f"{abs(float(balance)):.2f}"
            expected_balance = f"{abs(float(expected)):.2f}"
            if abs(float(balance)) > abs(float(expected)) and noisy_balance.endswith(expected_balance):
                txn["balance"] = expected
                balance = expected
        if expected is not None and (balance is None or abs(float(balance) - expected) <= tolerance):
            txn["balance"] = expected
            balance = expected
        if balance is not None:
            previous = float(balance)
    return txns


def reconcile_icbc_statement_totals(text: str, txns: list[dict]) -> list[dict]:
    rows = re.findall(
        r"本页支出算术合计：\s*([\d,]+\.\d{2})\s*(?:-\s*)?"
        r"本页收入算术合计：\s*([\d,]+\.\d{2})",
        text,
    )
    if not rows or not txns:
        return txns
    expected_expense = round(sum(float(a.replace(",", "")) for a, _ in rows), 2)
    expected_income = round(sum(float(b.replace(",", "")) for _, b in rows), 2)
    actual_expense = round(sum(float(x.get("expense") or 0) for x in txns), 2)
    actual_income = round(sum(float(x.get("income") or 0) for x in txns), 2)
    # 仅校正防伪扰码导致的微小尾差；较大差异继续暴露，避免掩盖漏行。
    for field, expected, actual, sign in (
        ("expense", expected_expense, actual_expense, -1),
        ("income", expected_income, actual_income, 1),
    ):
        residual = round(expected - actual, 2)
        if not residual or abs(residual) > 100:
            continue
        target = next((x for x in reversed(txns) if float(x.get(field) or 0) > 0), None)
        if target:
            corrected = round(float(target.get(field) or 0) + residual, 2)
            target[field] = corrected
            target["amount"] = round(sign * corrected, 2)
    return txns


def parse_boc_ocr_line(line: str, account: str, source: str) -> dict | None:
    text = normalize_money_fragments(line)
    # 修复中行文字层偶发丢失日期分隔符：2026-0424、202604-17、20251124。
    text = re.sub(
        r"^(20\d{2})-?(\d{2})-?(\d{2})(?=\s+\d{1,2}\s*:)",
        r"\1-\2-\3",
        text,
    )
    match = re.match(
        r"^(?P<date>20\d{2}[-/.]\d{1,2}[-/.]\d{1,2})\s+"
        r"(?P<time>\d{1,2}\s*:\s*\d{2}\s*:\s*\d{2})\s+人民币\s+"
        r"(?P<amount>[+-]?(?:\d{1,3}(?:[,，]\d{3})+|\d+)(?:\.\d{1,2})?)\s+"
        r"(?P<balance>[+-]?(?:\d{1,3}(?:[,，]\d{3})+|\d+)(?:\.\d{1,2})?)\s+"
        r"(?P<rest>.+)$",
        text,
    )
    if not match:
        return None
    amount = money_to_float(match.group("amount"))
    balance = money_to_float(match.group("balance"))
    date = parse_date(match.group("date"))
    if amount is None or not date:
        return None
    rest = normalize_text(match.group("rest"))
    parts = [part for part in re.split(r"\s+", rest) if part]
    counterparty = infer_counterparty(parts)
    for idx, part in enumerate(parts):
        if re.fullmatch(r"[0-9A-Z-]{8,}", part):
            before = [p for p in parts[max(0, idx - 3) : idx] if not re.search(r"手机银行|网上银行|银企对接|跨行转账|转账收入|贷款还款|网上快捷", p)]
            if before:
                counterparty = before[-1]
                break
    return make_txn(
        date=date,
        account=account,
        counterparty=counterparty,
        summary=rest,
        income=amount if amount > 0 else None,
        expense=abs(amount) if amount < 0 else None,
        amount=amount,
        balance=balance,
        source=source,
        raw_key=line,
    )


def parse_minsheng_statement_line(line: str, account: str, source: str) -> dict | None:
    if "|" in line:
        return None
    match = re.match(
        r"^(?P<date>20\d{2}/\d{2}/\d{2})\s+(?P<body>.*?)\s+"
        r"(?P<debit>[\d,]+\.\d{2})\s+(?P<credit>[\d,]+\.\d{2})\s+"
        r"(?P<balance>[\d,]+\.\d{2})\s+(?P<flow>[A-Za-z0-9]{8,})(?:\s+(?P<rest>.*))?$",
        normalize_text(line),
    )
    if not match:
        return None
    debit = money_to_float(match.group("debit"))
    credit = money_to_float(match.group("credit"))
    balance = money_to_float(match.group("balance"))
    if not debit and not credit:
        return None
    body = normalize_text(match.group("body"))
    rest = normalize_text(match.group("rest"))
    counterparty = rest.split("/", 1)[0].strip() if rest else ""
    return make_txn(
        date=parse_date(match.group("date")),
        account=account,
        counterparty=clean_counterparty_text(counterparty or "中国民生银行"),
        summary=body,
        income=credit,
        expense=debit,
        amount=None,
        balance=balance,
        source=source,
        raw_key=line,
    )


def parse_citic_statement_line(line: str, account: str, source: str, previous_balance: float | None) -> tuple[dict | None, float | None]:
    match = re.match(
        r"^(?P<date>20\d{6})\s+(?P<seq>\d{6})\s+RMB\s+"
        r"(?P<amount>[\d,]+\.\d{1,2})\s+RMB\s+(?P<balance>[\d,]+\.\d{1,2})\s+"
        r"(?P<rest>.+?)\s+(?:N|Y)$",
        normalize_text(line),
    )
    if not match:
        return None, previous_balance
    amount = abs(float(money_to_float(match.group("amount")) or 0))
    balance = money_to_float(match.group("balance"))
    if not amount or balance is None:
        return None, previous_balance
    rest = normalize_text(match.group("rest"))
    effect = float(balance) - float(previous_balance) if previous_balance is not None else None
    tolerance = max(0.05, amount * 0.0001)
    if effect is not None and abs(abs(effect) - amount) <= tolerance:
        direction = "收入" if effect > 0 else "支出"
    else:
        direction = detect_direction(rest, None)
    parts = rest.split()
    counterparty = parts[-1] if len(parts) > 1 else infer_counterparty(parts)
    txn = make_txn(
        date=parse_date(match.group("date")),
        account=account,
        counterparty=counterparty,
        summary=rest,
        income=amount if direction == "收入" else None,
        expense=amount if direction == "支出" else None,
        amount=amount if direction == "收入" else -amount,
        balance=balance,
        source=source,
        raw_key=line,
    )
    return txn, balance


def dedupe_transactions(txns: list[dict]) -> list[dict]:
    seen = set()
    unique = []
    for txn in txns:
        key = transaction_fingerprint(txn)
        if key in seen:
            continue
        seen.add(key)
        unique.append(txn)
    return unique


def transaction_fingerprint(txn: dict) -> tuple:
    account = re.sub(r"\s+", "", normalize_text(txn.get("account")))
    date = normalize_text(txn.get("date"))
    income = round(float(txn.get("income") or 0), 2)
    expense = round(float(txn.get("expense") or 0), 2)
    other_amount = round(float(txn.get("other_amount") or 0), 2)
    balance = txn.get("balance")
    raw_key = normalize_text(txn.get("raw_key"))
    time_match = re.search(r"(?:20\d{2}[-/.]?\d{2}[-/.]?\d{2})(?:[ T]+|\s*\|\s*)(\d{2}:?\d{2}:?\d{2})", raw_key)
    event_time = re.sub(r"\D", "", time_match.group(1)) if time_match else ""
    first_cell = normalize_text(raw_key.split("|", 1)[0]) if "|" in raw_key else ""
    event_id = re.sub(r"\D", "", first_cell) if first_cell else ""
    counterparty = re.sub(r"\s+", "", normalize_text(txn.get("counterparty")))
    if balance is not None:
        return (
            "balance",
            account,
            date,
            event_time,
            event_id,
            income,
            expense,
            round(float(balance), 2),
        )
    transaction_match = re.search(r"(?<!\d)(?:\d[\d\s]{14,}\d)(?!\d)", raw_key)
    transaction_id = re.sub(r"\s+", "", transaction_match.group(0)) if transaction_match else ""
    if len(transaction_id) >= 12:
        return ("transaction_id", account, transaction_id, date, income, expense, other_amount)
    summary = re.sub(r"\s+", "", normalize_text(txn.get("summary")))
    return ("content", account, date, income, expense, other_amount, counterparty, summary)


def dedupe_across_files(transactions: list[dict]) -> tuple[list[dict], dict[str, int]]:
    file_groups: dict[str, list[dict]] = {}
    for txn in transactions:
        source_file = normalize_text(txn.get("source")).split(" / ", 1)[0] or "未识别文件"
        file_groups.setdefault(source_file, []).append(txn)

    max_occurrences: dict[tuple, int] = {}
    duplicate_counts: dict[str, int] = defaultdict(int)
    unique: list[dict] = []
    for source_file, file_txns in file_groups.items():
        local_occurrences: dict[tuple, int] = defaultdict(int)
        for txn in file_txns:
            fingerprint = transaction_fingerprint(txn)
            local_occurrences[fingerprint] += 1
            occurrence = local_occurrences[fingerprint]
            if occurrence <= max_occurrences.get(fingerprint, 0):
                duplicate_counts[source_file] += 1
                continue
            unique.append(txn)
        for fingerprint, count in local_occurrences.items():
            max_occurrences[fingerprint] = max(max_occurrences.get(fingerprint, 0), count)
    return unique, dict(duplicate_counts)


def extract_text_transactions(text: str, source: str, should_cancel=None) -> list[dict]:
    txns: list[dict] = []
    default_year = None
    account = infer_account_from_text(text, source)
    shunde_rcb_mode = "广东顺德农村商业银行" in text and "账户/卡明细信息" in text and "存入/支取" in text
    if shunde_rcb_mode:
        return dedupe_transactions(parse_shunde_rcb_personal_text(text, account, source))
    wechat_mode = "微信支付交易明细证明" in text and "|" in text
    alipay_mode = "支付宝支付科技有限公司" in text and "收/支" in text
    minsheng_mode = "中国民生银行" in text and "借方发生额" in text and "贷方发生额" in text
    citic_mode = "账户交易明细" in text and "账户序号" in text and "中信银行" in source
    boc_personal_mode = "中国银行交易流水明细清单" in text and "借记卡号" in text
    previous_citic_balance = None
    pipe_table_mode = bool(
        re.search(
            r"(?:交易时\s*间|交易时间|交易日期|会计日期|记账日|流水号|收[/／]?支)\s*\|.*(?:交易金额|金额|收入|支出|存入|转出金额|转入金额|借方发\s*生额|贷方发\s*生额|借方发生额|贷方发生额|账户余额|交易余额|借贷标志)",
            text,
        )
    ) or bool(
        re.search(r"交易日期[^\n|]*\|[^\n]*(?:交易金额|Transaction Amount)", text)
        or re.search(r"(?m)^[^\n]*(?:交易日期|交易时间|交易流水)[^\n]*\|[^\n]*(?:交易金额|借方|贷方|余额)", text)
    ) or alipay_mode
    # 中国银行导出的文本同时含有视觉文本和表格文本；只解析带竖线的表格行，
    # 避免同一笔交易被两套文本层重复计入。
    if "中国银行交易流水明细清单" in text and re.search(
        r"20\d{2}-\d{2}-\d{2}\s*\|\s*\d{2}:", text
    ):
        pipe_table_mode = True
    saw_table_header = False
    in_table = False
    pending_grcb_prefix = ""
    for line_no, raw in enumerate(text.splitlines(), start=1):
        if line_no % 50 == 0:
            raise_if_cancelled(should_cancel)
        line = normalize_text(raw)
        if len(line) < 8:
            continue
        # 中行 PDF 每页还附带一套逐字段文本层。该文本层中的余额行会被通用
        # 日期/金额规则误判；个人流水只接收以完整日期和时间开头的视觉表格行。
        if boc_personal_mode and not pipe_table_mode:
            structured = parse_boc_ocr_line(line, account, source)
            if structured:
                txns.append(structured)
            continue
        if pipe_table_mode:
            if alipay_mode:
                structured = parse_alipay_table_line(line, account, source)
            elif wechat_mode:
                structured = parse_wechat_table_line(line, source, account)
            else:
                if is_grcb_broken_prefix_line(line):
                    pending_grcb_prefix = line
                    continue
                if pending_grcb_prefix:
                    merged = merge_grcb_broken_line(pending_grcb_prefix, line)
                    if merged:
                        line = merged
                        pending_grcb_prefix = ""
                structured = parse_bank_pipe_table_line(line, account, source)
            if structured:
                txns.append(structured)
            continue
        if is_table_header(line):
            saw_table_header = True
            in_table = True
            continue
        if saw_table_header and is_table_footer(line):
            in_table = False
            continue
        if saw_table_header and not in_table:
            continue
        if is_non_transaction_line(line):
            continue
        if citic_mode:
            citic_txn, citic_balance = parse_citic_statement_line(line, account, source, previous_citic_balance)
            if citic_txn:
                txns.append(citic_txn)
                previous_citic_balance = citic_balance
                continue
        structured = parse_wechat_table_line(line, source, account)
        if wechat_mode:
            if structured:
                txns.append(structured)
            continue
        structured = (
            structured
            or (parse_minsheng_statement_line(line, account, source) if minsheng_mode else None)
            or parse_boc_ocr_line(line, account, source)
            or parse_abc_line(line, account, source)
        )
        if structured:
            txns.append(structured)
            continue
        if re.search(r"\d{10,}", line) and not re.search(r"[+-]?\s*\d[\d,，\s]*\.\s*\d{1,2}|[+-]\s*\d", line):
            continue
        date = parse_date(line, default_year)
        if not date:
            continue
        default_year = int(date[:4])
        signed_values = extract_signed_amounts_from_line(line)
        values = extract_amounts_from_line(line)
        values = [v for v in values if abs(v) >= 0.01 and str(int(abs(v))) not in date.replace("-", "")]
        if not values:
            continue
        direction = detect_direction(line, signed_values[0] if signed_values else values[0])
        income = expense = balance = amount = None
        if signed_values:
            amount = signed_values[0]
            try:
                signed_index = values.index(amount)
            except ValueError:
                signed_index = -1
            after_signed = values[signed_index + 1 :] if signed_index >= 0 else []
            balance_candidates = [v for v in after_signed if v >= 0 and abs(v) != abs(amount)]
            balance = balance_candidates[0] if balance_candidates else None
        elif len(values) >= 2:
            balance = values[-1]
            amount = values[-2] if abs(values[-2]) <= abs(balance) * 20 + 100000000 else values[0]
        else:
            amount = values[0]
        if amount < 0:
            expense = abs(amount)
        elif direction == "支出":
            expense = abs(amount)
        else:
            income = abs(amount)
        parts = re.split(r"\s+|\|", line)
        counterparty = infer_counterparty_after_amount(line) if signed_values else ""
        counterparty = counterparty or infer_counterparty(parts)
        txn = make_txn(
            date=date,
            account=account,
            counterparty=counterparty,
            summary=line,
            income=income,
            expense=expense,
            amount=amount,
            balance=balance,
            source=source,
            raw_key=line,
        )
        if txn:
            txns.append(txn)
    if "中国工商银行借记账户历史明细" in text:
        txns = repair_running_balances(txns)
        txns = reconcile_icbc_statement_totals(text, txns)
    return dedupe_transactions(txns)


def analyze_file(path: Path, progress=None, passwords: list[str] | None = None, should_cancel=None) -> dict:
    raise_if_cancelled(should_cancel)
    passwords = passwords or [""]
    suffix = path.suffix.lower()
    job_dir = RESULTS / path.stem
    if job_dir.exists():
        shutil.rmtree(job_dir)
    job_dir.mkdir(parents=True, exist_ok=True)
    if suffix in EXCEL_EXTS:
        if progress:
            progress(12, "读取表格文件")
        txns = extract_excel(path, progress=progress, passwords=passwords, should_cancel=should_cancel)
        text_length = 0
        source_mode = "Excel/CSV表格"
        info = extract_statement_info("", path.name)
    elif suffix in PDF_EXTS:
        if progress:
            progress(10, "检查 PDF 文字层")
        text, used_password = extract_pdf_text(path, passwords=passwords, should_cancel=should_cancel, progress=progress)
        source_mode = "PDF文字层"
        if len(text.strip()) < 120:
            text = ocr_file(
                path,
                job_dir,
                progress=progress,
                passwords=[used_password] if used_password else passwords,
                should_cancel=should_cancel,
            )
            source_mode = "OCR识别"
        if progress:
            progress(88, "解析流水文本")
        if is_corporate_query_statement(text, path.name):
            txns, _ = extract_corporate_query_pdf(
                path,
                passwords=[used_password] if used_password else passwords,
                should_cancel=should_cancel,
            )
            source_mode = "PDF文字层-企业交易查询"
        elif is_abc_account_detail_statement(text, path.name):
            txns, _ = extract_abc_account_detail_pdf(
                path,
                passwords=[used_password] if used_password else passwords,
                should_cancel=should_cancel,
            )
            source_mode = "PDF文字层-农行账户明细"
        else:
            txns = extract_text_transactions(text, path.name, should_cancel=should_cancel)
        text_length = len(text)
        info = extract_statement_info(text, path.name)
        (job_dir / "extracted.txt").write_text(text, encoding="utf-8")
    elif suffix in IMAGE_EXTS:
        text = ocr_file(path, job_dir, progress=progress, passwords=passwords, should_cancel=should_cancel)
        source_mode = "OCR识别"
        if progress:
            progress(88, "解析图片流水")
        txns = extract_text_transactions(text, path.name, should_cancel=should_cancel)
        text_length = len(text)
        info = extract_statement_info(text, path.name)
        (job_dir / "extracted.txt").write_text(text, encoding="utf-8")
    else:
        raise ValueError("目前支持 PDF、Excel、CSV 和图片文件。")
    accounts = sorted({txn.get("account") for txn in txns if txn.get("account")})

    return {
        "filename": path.name,
        "source_mode": source_mode,
        "transaction_count": len(txns),
        "text_length": text_length,
        "info": info,
        "accounts": accounts,
        "transactions": txns,
    }


def aggregate_transactions(transactions: list[dict]) -> dict:
    monthly: dict[str, dict] = defaultdict(lambda: {"income": 0.0, "expense": 0.0, "valid_income": 0.0, "interest_income": 0.0, "count": 0, "valid_count": 0, "interest_count": 0})
    counterparties: dict[str, dict] = defaultdict(lambda: {"income": 0.0, "expense": 0.0, "count": 0})
    for txn in transactions:
        month = txn.get("month") or "未识别日期"
        income = float(txn.get("income") or 0)
        expense = float(txn.get("expense") or 0)
        monthly[month]["income"] += income
        monthly[month]["expense"] += expense
        monthly[month]["count"] += 1
        if income and not txn.get("default_excluded"):
            monthly[month]["valid_income"] += income
            monthly[month]["valid_count"] += 1
        if txn.get("category") == "银行结息":
            monthly[month]["interest_income"] += income
            monthly[month]["interest_count"] += 1
        cp = txn.get("counterparty") or "未识别"
        counterparties[cp]["income"] += income
        counterparties[cp]["expense"] += expense
        counterparties[cp]["count"] += 1
    return {
        "monthly": [
            {"month": k, **{kk: round(vv, 2) if isinstance(vv, float) else vv for kk, vv in v.items()}}
            for k, v in sorted(monthly.items())
        ],
        "counterparties": [
            {"name": k, **{kk: round(vv, 2) if isinstance(vv, float) else vv for kk, vv in v.items()}}
            for k, v in sorted(counterparties.items(), key=lambda item: item[1]["income"], reverse=True)
        ][:200],
    }


def safe_filename(value: str) -> str:
    text = normalize_text(value) or "流水统计报告"
    text = re.sub(r"[\\/:*?\"<>|]+", "_", text)
    return text[:80] or "流水统计报告"


def export_cell(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value)
    amount = money_to_float(text)
    if amount is not None and re.fullmatch(r"[-+]?[\d,，]+(?:\.\d+)?", text.strip()):
        return round(amount, 2)
    return text


def normalize_export_blocks(data: dict) -> list[dict]:
    blocks = []
    for idx, block in enumerate(data.get("blocks") or [], start=1):
        headers = [normalize_text(item) for item in block.get("headers") or []]
        rows = block.get("rows") or []
        if not headers or not isinstance(rows, list):
            continue
        width = len(headers)
        normalized_rows = []
        for row in rows:
            if not isinstance(row, list):
                continue
            values = [export_cell(value) for value in row[:width]]
            if len(values) < width:
                values.extend([""] * (width - len(values)))
            normalized_rows.append(values)
        blocks.append({
            "title": normalize_text(block.get("title")) or f"表{idx}",
            "headers": headers,
            "rows": normalized_rows,
        })
    if not blocks:
        raise ValueError("没有可导出的统计结果，请先完成流水统计。")
    return blocks


def build_flow_xlsx_export(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    header_fill = PatternFill("solid", fgColor="EAF1F8")
    summary_fill = PatternFill("solid", fgColor="EAF7EF")
    header_font = Font(bold=True, color="1F334A")

    def unique_sheet_name(name: str) -> str:
        base = re.sub(r"[\[\]:*?/\\]", "", name)[:28] or "统计"
        candidate = base
        idx = 1
        while candidate in wb.sheetnames:
            suffix = f"_{idx}"
            candidate = f"{base[:31-len(suffix)]}{suffix}"
            idx += 1
        return candidate

    for block in normalize_export_blocks(data):
        ws = wb.create_sheet(unique_sheet_name(block["title"]))
        ws.append(block["headers"])
        for row in block["rows"]:
            ws.append(row)
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            label = normalize_text(row[0].value if row else "")
            is_summary = bool(re.search(r"合计|月均|小计", label))
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
                if is_summary:
                    cell.fill = summary_fill
                    cell.font = Font(bold=True)
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = max(
                len(str(ws.cell(row=row, column=col).value or ""))
                for row in range(1, min(ws.max_row, 100) + 1)
            )
            ws.column_dimensions[letter].width = min(max(max_len + 3, 12), 36)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_flow_pdf_export(data: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    font_name = "Helvetica"
    for font_path in [
        "/System/Library/Fonts/STHeiti Light.ttc",
        "/System/Library/Fonts/Hiragino Sans GB.ttc",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
    ]:
        try:
            pdfmetrics.registerFont(TTFont("FlowExportCN", font_path))
            font_name = "FlowExportCN"
            break
        except Exception:
            continue

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("FlowTitle", parent=styles["Title"], fontName=font_name, fontSize=18, leading=24, alignment=1)
    note_style = ParagraphStyle("FlowNote", parent=styles["BodyText"], fontName=font_name, fontSize=8, leading=11, alignment=1)
    h_style = ParagraphStyle("FlowHeading", parent=styles["Heading2"], fontName=font_name, fontSize=12, leading=16, spaceBefore=8, spaceAfter=4)
    cell_style = ParagraphStyle("FlowCell", parent=styles["BodyText"], fontName=font_name, fontSize=6.5, leading=8)

    bio = BytesIO()
    page = landscape(A4)
    doc = SimpleDocTemplate(bio, pagesize=page, leftMargin=9 * mm, rightMargin=9 * mm, topMargin=9 * mm, bottomMargin=9 * mm)
    title = normalize_text(data.get("title")) or "流水统计报告"
    generated = normalize_text(data.get("generated_at")) or datetime.now().strftime("%Y-%m-%d %H:%M")
    filters = normalize_text(data.get("filters_text")) or "当前页面筛选口径"
    story = [
        Paragraph(html.escape(title), title_style),
        Paragraph(html.escape(f"生成时间：{generated}　筛选口径：{filters}"), note_style),
        Spacer(1, 4 * mm),
    ]
    usable_width = page[0] - 18 * mm
    for block in normalize_export_blocks(data):
        story.append(Paragraph(html.escape(block["title"]), h_style))
        table_data = [[Paragraph(html.escape(str(h)), cell_style) for h in block["headers"]]]
        for row in block["rows"]:
            table_data.append([
                Paragraph(html.escape(money(value) if isinstance(value, (int, float)) else str(value or "")), cell_style)
                for value in row
            ])
        col_count = max(1, len(block["headers"]))
        table = Table(table_data, colWidths=[usable_width / col_count] * col_count, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAF1F8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1F334A")),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2.5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2.5),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ]))
        story.extend([table, Spacer(1, 3 * mm)])
    doc.build(story)
    return bio.getvalue()


def build_flow_export(data: dict, fmt: str) -> tuple[bytes, str, str]:
    fmt = (fmt or "").lower()
    title = safe_filename(normalize_text(data.get("title")) or "流水统计报告")
    if fmt == "pdf":
        return build_flow_pdf_export(data), f"{title}.pdf", "application/pdf"
    if fmt in {"excel", "xlsx"}:
        return build_flow_xlsx_export(data), f"{title}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    raise ValueError("下载格式不支持。")


def build_flow_response(file_results: list[dict], transactions: list[dict]) -> dict:
    unique_transactions, duplicate_counts = dedupe_across_files(transactions)
    for item in file_results:
        filename = normalize_text(item.get("filename"))
        duplicate_count = duplicate_counts.get(filename, 0)
        item["duplicate_transaction_count"] = duplicate_count
        item["unique_transaction_count"] = max(0, int(item.get("transaction_count") or 0) - duplicate_count)
    return {
        "files": file_results,
        "transactions": unique_transactions,
        "aggregate": aggregate_transactions(unique_transactions),
        "duplicate_transaction_count": sum(duplicate_counts.values()),
        "default_exclude_keywords": DEFAULT_EXCLUDE_KEYWORDS,
    }


def run_flow_job(job_id: str, files: list[Path], passwords: list[str] | None = None, append_existing: bool = False) -> None:
    try:
        passwords = passwords or [""]
        job = get_job(job_id)
        all_results = list(job.get("partial_files_results") or []) if append_existing else []
        all_txns: list[dict] = list(job.get("partial_transactions") or []) if append_existing else []
        pending_password_files: list[str] = []
        password_messages: dict[str, str] = {}
        total = len(files) or 1

        def should_cancel() -> bool:
            return is_job_cancelled(job_id)

        for idx, path in enumerate(files, start=1):
            raise_if_cancelled(should_cancel)
            base = int((idx - 1) * 100 / total)
            span = 100 / total

            def progress(file_pct, message, idx=idx, path=path):
                raise_if_cancelled(should_cancel)
                set_job(
                    job_id,
                    status="running",
                    percent=min(99, int(base + span * file_pct / 100)),
                    message=f"第 {idx}/{total} 个文件：{path.name} - {message}",
                    current_file=idx,
                    total_files=total,
                )

            progress(0, "开始处理")
            try:
                result = analyze_file(path, progress=progress, passwords=passwords, should_cancel=should_cancel)
            except PasswordRequiredError as exc:
                pending_password_files.append(str(path))
                password_messages[path.name] = str(exc)
                set_job(
                    job_id,
                    status="running",
                    percent=min(99, int(base + span)),
                    message=f"第 {idx}/{total} 个文件：{path.name} 需要密码，已先跳过并继续处理其他文件",
                    current_file=idx,
                    total_files=total,
                )
                continue
            all_results.append({k: v for k, v in result.items() if k != "transactions"})
            all_txns.extend(result["transactions"])
        raise_if_cancelled(should_cancel)
        response = build_flow_response(all_results, all_txns)
        if pending_password_files:
            names = [Path(item).name for item in pending_password_files]
            preview = "、".join(names[:6]) + (" 等" if len(names) > 6 else "")
            set_job(
                job_id,
                status="needs_password",
                percent=99 if all_results else get_job(job_id).get("percent", 0),
                message=f"{len(names)} 个文件需要密码：{preview}。已完成的文件不会重复识别，输入密码后只补跑这些文件。",
                password_file=names[0],
                password_files=names,
                password_messages=password_messages,
                pending_password_files=pending_password_files,
                partial_files_results=all_results,
                partial_transactions=all_txns,
                result=response,
            )
            return
        set_job(
            job_id,
            status="done",
            percent=100,
            message="流水统计完成",
            result=response,
            pending_password_files=[],
            password_files=[],
            password_file="",
        )
    except JobCancelledError:
        set_job(
            job_id,
            status="cancelled",
            percent=get_job(job_id).get("percent", 0),
            message="已停止统计",
            cancel_requested=False,
        )
    except Exception as exc:
        set_job(job_id, status="error", percent=100, message=str(exc), error=str(exc), trace=traceback.format_exc())


INDEX = r"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>流水统计平台</title>
<style>
:root{--ink:#162033;--muted:#677485;--line:#d7dee8;--panel:#fff;--soft:#f4f7f9;--nav:#20252d;--blue:#2d6f9f;--green:#247550;--amber:#8a6118;--red:#b42318;--shadow:0 12px 24px rgba(25,34,45,.08)}
*{box-sizing:border-box}body{margin:0;background:#eef2f5;color:var(--ink);font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft YaHei",sans-serif;line-height:1.5}button,input,textarea,select{font:inherit}.shell{display:grid;grid-template-columns:360px minmax(0,1fr);min-height:100vh}.side{background:var(--nav);color:#fff;padding:22px;display:flex;flex-direction:column;gap:16px}.main{padding:22px;min-width:0}.brand h1{font-size:22px;margin:0 0 5px}.brand p{margin:0;color:#c7d1dc;font-size:13px}.box{border:1px solid rgba(255,255,255,.18);background:rgba(255,255,255,.06);border-radius:8px;padding:14px}.drop{border:1px dashed rgba(255,255,255,.36);border-radius:8px;min-height:112px;display:grid;place-items:center;text-align:center;padding:14px;color:#dce6f1}.drop strong{display:block;color:#fff;margin-bottom:3px}.drop.drag{background:rgba(80,150,180,.18);border-color:#9dd7e8}.native-file{width:100%;margin-top:12px;color:#dce6f1}.file-name{font-size:13px;color:#c5d1df;margin-top:10px;word-break:break-all}.btn-row{display:flex;gap:8px;flex-wrap:wrap;margin-top:12px}button{border:0;border-radius:8px;min-height:38px;padding:9px 13px;cursor:pointer;background:#e8edf3;color:#172033}button.primary{background:#2d7d9a;color:#fff}button.danger{background:#b42318;color:#fff}button.ghost{background:transparent;color:#dce6f1;border:1px solid rgba(255,255,255,.24)}button:disabled{opacity:.55;cursor:not-allowed}.status{font-size:13px;color:#dce6f1;white-space:pre-wrap;min-height:42px;margin-top:12px}.progress{height:10px;background:rgba(255,255,255,.16);border-radius:999px;overflow:hidden;margin-top:10px}.progress span{display:block;height:100%;width:0;background:#7fd1a6;transition:width .25s ease}.field{display:grid;gap:6px;margin-bottom:12px}.field label{font-size:13px;color:#d7e0ea}.field textarea,.field input,.field select{width:100%;border:1px solid rgba(255,255,255,.24);background:rgba(255,255,255,.08);color:#fff;border-radius:8px;padding:9px;min-height:38px}.field textarea{min-height:80px;resize:vertical}.check{display:flex;gap:8px;align-items:flex-start;color:#d7e0ea;font-size:13px}.check input{margin-top:3px}.metrics{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin-bottom:16px}.metric{background:var(--panel);border:1px solid var(--line);border-radius:8px;padding:14px;box-shadow:var(--shadow)}.metric span{display:block;color:var(--muted);font-size:13px;margin-bottom:8px}.metric strong{font-size:24px;line-height:1.1;word-break:break-word}.metric small{display:block;color:var(--muted);font-size:12px;margin-top:7px}.panel{background:var(--panel);border:1px solid var(--line);border-radius:8px;box-shadow:var(--shadow);padding:16px;margin-bottom:16px;overflow:hidden}.toolbar{display:flex;justify-content:space-between;align-items:center;gap:12px;margin-bottom:12px}.toolbar h2{font-size:18px;margin:0}.actions{display:flex;gap:8px;flex-wrap:wrap;justify-content:flex-end}.actions button{background:#eef4f7;color:#245d72}.empty{color:var(--muted);padding:26px;text-align:center;background:var(--soft);border:1px dashed var(--line);border-radius:8px}.tabs{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px}.tab{border:1px solid var(--line);background:#f3f6f9;color:#405065}.tab.active{background:#25313d;color:#fff;border-color:#25313d}.table-wrap{overflow:auto}.file-block{border-top:2px solid var(--line);padding-top:14px;margin-top:18px}.file-block:first-child{border-top:0;margin-top:0;padding-top:0}.file-title{display:flex;gap:10px;align-items:baseline;justify-content:space-between;flex-wrap:wrap;margin:0 0 8px}.file-title h3{font-size:16px;margin:0}.file-title span{color:var(--muted);font-size:13px}table{border-collapse:collapse;width:100%;min-width:860px;margin:8px 0 12px}th,td{border:1px solid #cfd7e3;padding:8px;text-align:left;font-size:14px;vertical-align:top}th{background:#f3f6f9;color:#405065}.summary-row th,.summary-row td{background:#eaf1f0;font-weight:700}.pill{display:inline-flex;border-radius:999px;padding:3px 8px;font-size:12px;font-weight:650;background:#eef4f7;color:var(--blue)}.pill.off{background:#f7eee8;color:#8d4f2c}.debug{font-size:12px;color:var(--muted);background:#f7f9fb;border:1px solid var(--line);border-radius:8px;padding:10px;max-height:130px;overflow:auto;white-space:pre-wrap}.cp-list{display:flex;gap:8px;flex-wrap:wrap;max-height:130px;overflow:auto}.cp{border:1px solid var(--line);border-radius:999px;background:#f7f9fb;padding:6px 10px;font-size:13px;cursor:pointer}.cp.active{background:#fdebea;border-color:#e9b5af;color:var(--red)}@media(max-width:980px){.shell{grid-template-columns:1fr}.metrics{grid-template-columns:1fr 1fr}.side{min-height:auto}}@media(max-width:620px){.metrics{grid-template-columns:1fr}.main{padding:14px}.side{padding:16px}}@media print{.side,.actions,.debug,.tabs,.cp-list{display:none}.shell{display:block}.main{padding:0}.panel,.metric{box-shadow:none;border:0}.metrics{display:none}}
.risk-cards{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;margin:8px 0 14px}.risk-card{border:1px solid var(--line);border-radius:8px;padding:12px;background:#f7f9fb}.risk-card span{display:block;color:var(--muted);font-size:12px}.risk-card strong{display:block;font-size:20px;margin-top:5px}.risk-high{color:var(--red)}.risk-medium{color:var(--amber)}.risk-low{color:var(--green)}.risk-note{padding:11px 12px;border-left:4px solid var(--blue);background:#eef5f8;border-radius:6px;margin:8px 0 14px}.pill.high{background:#fdebea;color:var(--red)}.pill.medium{background:#fff4dd;color:var(--amber)}.pill.low{background:#eaf7f0;color:var(--green)}@media(max-width:980px){.risk-cards{grid-template-columns:1fr 1fr}}@media(max-width:620px){.risk-cards{grid-template-columns:1fr}}
</style></head><body><div class="shell"><aside class="side"><div class="brand"><h1>流水统计平台</h1><p>PDF、Excel、图片流水本机统计</p></div>
<section class="box"><form id="form"><div class="drop" id="drop"><span><strong>拖入流水文件</strong><small>PDF、Excel、CSV 或图片</small></span></div><input id="file" class="native-file" name="file" type="file" accept=".pdf,.xlsx,.xls,.csv,image/*" multiple required><div id="fileName" class="file-name">尚未选择文件</div><div class="field" style="margin-top:12px"><label>文件密码</label><textarea id="passwords" name="passwords" autocomplete="off" placeholder="每行一个密码，也可写：工行 784409&#10;中行=515545"></textarea></div><div class="btn-row"><button id="submitBtn" class="primary">上传并统计</button><button id="stopBtn" class="danger" type="button" style="display:none">停止统计</button><button id="resumeBtn" class="primary" type="button" style="display:none">输入密码后继续</button><button id="resetBtn" class="ghost" type="button">清空</button></div></form><div class="progress"><span id="progressBar"></span></div><div id="status" class="status">请选择流水文件。带水印扫描件会自动做去水印增强识别。</div></section>
<section class="box"><div class="field"><label>剔除对象</label><textarea id="excludeNames" placeholder="多个对象用逗号或换行分隔"></textarea></div><label class="check"><input id="excludeDefault" type="checkbox" checked>剔除常见非经营流水</label><label class="check"><input id="excludeFinancial" type="checkbox">剔除银行/金融机构交易对手</label><div class="field" style="margin-top:12px"><label>统计方向</label><select id="directionMode"><option value="income">收入流水</option><option value="all">收入和支出</option></select></div><button id="applyFilter" type="button">重新统计</button></section>
<section class="box"><div class="brand"><h1 style="font-size:16px">银行风控检测</h1><p>按经营范围自动建立行业模型</p></div><div class="field"><label>客户类型</label><select id="borrowerType"><option value="enterprise">企业</option><option value="individual">个人经营</option></select></div><div class="field"><label>客户/企业名称</label><input id="subjectName" placeholder="用于识别本人及关联方互转"></div><div class="field"><label>营业执照经营范围</label><textarea id="businessScope" placeholder="粘贴营业执照上的经营范围"></textarea></div><div class="field"><label>实际经营情况</label><textarea id="businessDescription" placeholder="例如：铝型材加工，向工厂采购原料，主要对公销售，回款周期约30天"></textarea></div><div class="field"><label>关联方名单</label><textarea id="relatedParties" placeholder="股东、配偶、关联公司等，逗号或换行分隔"></textarea></div><div class="field"><label>拟贷款金额（元）</label><input id="loanAmount" type="number" min="0" step="10000" placeholder="可不填"></div><div class="field"><label>贷款期限（月）</label><input id="loanTerm" type="number" min="1" max="360" value="12"></div><button id="riskAnalyzeBtn" class="primary" type="button">生成银行风控检测</button><div class="note" style="color:#c7d1dc;margin-top:9px">结果用于贷前筛查和人工复核，不等同于银行最终审批结论。</div></section>
</aside><main class="main"><section class="metrics"><div class="metric"><span>有效流水合计</span><strong id="mValid">-</strong><small id="mValidSub">等待统计</small></div><div class="metric"><span>月均有效流水</span><strong id="mAvg">-</strong><small id="mAvgSub">等待统计</small></div><div class="metric"><span>有效笔数</span><strong id="mCount">-</strong><small id="mCountSub">等待统计</small></div><div class="metric"><span>剔除笔数</span><strong id="mExcluded">-</strong><small id="mExcludedSub">等待统计</small></div></section>
<section class="panel"><div class="toolbar"><h2>筛选对象</h2><div class="actions"><button id="printBtn" type="button">打印</button><button id="copyBtn" type="button">复制汇总</button></div></div><div id="counterparties" class="cp-list"><span class="empty">上传后显示交易对象。</span></div></section>
<section class="panel"><div class="toolbar"><h2>流水统计</h2><div class="actions"><button id="downloadPdfBtn" type="button" disabled>下载PDF</button><button id="downloadExcelBtn" type="button" disabled>下载Excel</button></div></div><div class="tabs"><button class="tab active" data-tab="monthly">全部汇总</button><button class="tab" data-tab="risk">银行风控检测</button><button class="tab" data-tab="files">单文件统计</button><button class="tab" data-tab="account">同账户汇总</button><button class="tab" data-tab="details">交易明细</button><button class="tab" data-tab="summary">汇总栏</button></div><div id="result"><div class="empty">上传流水后，这里会按月份列出有效流水，并在底部显示汇总。</div></div></section>
<section class="panel"><div class="toolbar"><h2>识别诊断</h2></div><div id="debug" class="debug">等待统计。</div></section></main></div>
<script>
const form=document.getElementById('form'),fileInput=document.getElementById('file'),drop=document.getElementById('drop'),fileName=document.getElementById('fileName'),statusEl=document.getElementById('status'),progressBar=document.getElementById('progressBar'),result=document.getElementById('result'),debug=document.getElementById('debug'),submitBtn=document.getElementById('submitBtn'),stopBtn=document.getElementById('stopBtn'),resumeBtn=document.getElementById('resumeBtn'),passwordsInput=document.getElementById('passwords'),excludeNames=document.getElementById('excludeNames'),excludeDefault=document.getElementById('excludeDefault'),excludeFinancial=document.getElementById('excludeFinancial'),directionMode=document.getElementById('directionMode'),counterparties=document.getElementById('counterparties'),borrowerType=document.getElementById('borrowerType'),subjectName=document.getElementById('subjectName'),businessScope=document.getElementById('businessScope'),businessDescription=document.getElementById('businessDescription'),relatedParties=document.getElementById('relatedParties'),loanAmount=document.getElementById('loanAmount'),loanTerm=document.getElementById('loanTerm'),riskAnalyzeBtn=document.getElementById('riskAnalyzeBtn');
let rawData=null,activeTab='monthly',clickedExcludes=new Set(),currentJobId=null;
const money=v=>Number(v||0).toLocaleString('zh-CN',{minimumFractionDigits:2,maximumFractionDigits:2});
const esc=v=>String(v??'').replace(/[&<>"']/g,ch=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[ch]));
function fileAccountText(file){const accounts=(file.accounts||[]).filter(Boolean);if(accounts.length)return accounts.join('，');const info=file.info||{};return info['账号']||'未识别账户';}
function fileInfoText(file){const info=file.info||{};const parts=['户名','账号','币种','起止日期','电子流水号'].map(k=>info[k]?`${k}：${info[k]}`:'').filter(Boolean);const account=fileAccountText(file);if(account&&account!=='未识别账户'&&!parts.some(x=>x.includes(account)))parts.unshift(`识别账号：${account}`);return parts.join('　')||'未识别到账户信息';}
function setProgress(p){progressBar.style.width=`${Math.max(0,Math.min(100,Number(p)||0))}%`;}
function setRunning(running){submitBtn.disabled=running;stopBtn.style.display=running?'inline-flex':'none';stopBtn.disabled=false;}
function setDownloadEnabled(enabled){['downloadPdfBtn','downloadExcelBtn'].forEach(id=>{const el=document.getElementById(id);if(el)el.disabled=!enabled;});}
function updateFile(){const files=[...fileInput.files];fileName.textContent=files.length?`${files.length} 个文件：`+files.map(f=>f.name).join('，'):'尚未选择文件';}
function excludeList(){const manual=excludeNames.value.split(/[,，\n\r;；]+/).map(x=>x.trim()).filter(Boolean);return [...new Set([...manual,...clickedExcludes])];}
function isFinancialCounterparty(txn){const hay=`${txn.counterparty||''} ${txn.summary||''}`;if(!hay.trim())return false;if(/银行结息|微信内部流水|手续费|贷款发放|贷款入账|放贷款|借款|还款|理财|基金|银证|保证金|融资租赁|小额贷款|小贷|消费金融|汽车金融|担保|保理|典当|信托|证券|保险|财付通|支付宝|网商银行|微众银行|京东金融|度小满|马上消费|招联金融|中邮消费|平安普惠/.test(hay))return true;if(/银行|信用社|农商行|农村商业银行|村镇银行|邮政储蓄|金融|证券|保险|基金|信托|租赁|保理|担保|贷款/.test(txn.counterparty||''))return true;return false;}
function isExcluded(txn){const names=excludeList();const hay=`${txn.counterparty||''} ${txn.summary||''}`;if(names.some(name=>hay.includes(name)))return true;if(excludeDefault.checked&&txn.default_excluded)return true;if(excludeFinancial.checked&&isFinancialCounterparty(txn))return true;return false;}
function isEffective(txn){if(isExcluded(txn))return false;if(directionMode.value==='income')return Number(txn.income||0)>0;return Number(txn.income||0)!==0||Number(txn.expense||0)!==0;}
function calc(scopeTxns){
  const txns=scopeTxns||((rawData&&rawData.transactions)||[]);
  const effective=txns.filter(isEffective);
  const excluded=txns.filter(x=>!isEffective(x));
  const months={};
  const accountMap={};
  const accountMonthlyMap={};
  const accountFileMap={};
  const ensureMonth=m=>months[m]??={month:m,income:0,expense:0,count:0,interest:0,interestCount:0};
  const ensureAccount=account=>accountMap[account]??={account,income:0,expense:0,count:0,months:new Set(),files:new Set(),interest:0,interestCount:0};
  const ensureAccountMonth=(account,month)=>accountMonthlyMap[`${account}||${month}`]??={account,month,income:0,expense:0,count:0,interest:0,interestCount:0};
  const ensureAccountFile=(account,file)=>accountFileMap[`${account}||${file}`]??={account,file,income:0,expense:0,count:0,months:new Set(),interest:0,interestCount:0};
  for(const t of txns){
    const m=t.month||'未识别日期';
    const account=t.account||t.source||'未识别账户';
    const source=t.source||'未识别文件';
    ensureMonth(m);
    if(t.category==='银行结息'){
      months[m].interest+=Number(t.income||0);
      months[m].interestCount++;
      const accountItem=ensureAccount(account);
      accountItem.months.add(m);
      accountItem.files.add(source);
      accountItem.interest+=Number(t.income||0);
      accountItem.interestCount++;
      const accountMonth=ensureAccountMonth(account,m);
      accountMonth.interest+=Number(t.income||0);
      accountMonth.interestCount++;
      const accountFile=ensureAccountFile(account,source);
      accountFile.months.add(m);
      accountFile.interest+=Number(t.income||0);
      accountFile.interestCount++;
    }
  }
  for(const t of effective){
    const m=t.month||'未识别日期';
    const account=t.account||t.source||'未识别账户';
    const source=t.source||'未识别文件';
    ensureMonth(m);
    months[m].income+=Number(t.income||0);
    months[m].expense+=Number(t.expense||0);
    months[m].count++;
    const accountItem=ensureAccount(account);
    accountItem.income+=Number(t.income||0);
    accountItem.expense+=Number(t.expense||0);
    accountItem.count++;
    accountItem.months.add(m);
    accountItem.files.add(source);
    const accountMonth=ensureAccountMonth(account,m);
    accountMonth.income+=Number(t.income||0);
    accountMonth.expense+=Number(t.expense||0);
    accountMonth.count++;
    const accountFile=ensureAccountFile(account,source);
    accountFile.income+=Number(t.income||0);
    accountFile.expense+=Number(t.expense||0);
    accountFile.count++;
    accountFile.months.add(m);
  }
  const monthly=Object.values(months).sort((a,b)=>a.month.localeCompare(b.month));
  const validMonthCount=monthly.filter(x=>x.count>0).length;
  const accountGroups=Object.values(accountMap).map(x=>({...x,files:[...x.files].filter(Boolean).sort(),monthCount:x.months.size,fileCount:x.files.size})).sort((a,b)=>b.income-a.income);
  const accountFiles=Object.values(accountFileMap).map(x=>({...x,monthCount:x.months.size})).sort((a,b)=>a.account.localeCompare(b.account)||a.file.localeCompare(b.file));
  const accountMonthly=Object.values(accountMonthlyMap).sort((a,b)=>a.account.localeCompare(b.account)||a.month.localeCompare(b.month));
  const totalIncome=effective.reduce((s,x)=>s+Number(x.income||0),0);
  const totalExpense=effective.reduce((s,x)=>s+Number(x.expense||0),0);
  const interestTotal=txns.filter(x=>x.category==='银行结息').reduce((s,x)=>s+Number(x.income||0),0);
  const interestCount=txns.filter(x=>x.category==='银行结息').length;
  return{txns,effective,excluded,monthly,validMonthCount,accountGroups,accountFiles,accountMonthly,totalIncome,totalExpense,interestTotal,interestCount,avg:validMonthCount?totalIncome/validMonthCount:0};
}
const INDUSTRY_MODELS=[
  {key:'agriculture',name:'农业/养殖/水产',pattern:/农业|种植|养殖|水产|渔业|农产品|饲料|生鲜/,topShare:.62,cv:1.05,fast:.72,round:.58,netMargin:.12,feature:'季节性和个人结算较常见，重点核验产销周期、饲料或农资采购与回款。'},
  {key:'construction',name:'建筑/工程/装修',pattern:/建筑|工程|施工|装修|装饰|安装|市政|园林|幕墙/,topShare:.68,cv:1.05,fast:.72,round:.62,netMargin:.10,feature:'项目制回款、客户集中和月度波动通常较高，重点核验合同、进度款和劳务材料支出。'},
  {key:'catering',name:'餐饮/食品零售',pattern:/餐饮|酒楼|饭店|食品|饮品|烘焙|便利店|超市|零售/,topShare:.32,cv:.58,fast:.55,round:.38,netMargin:.10,feature:'正常表现通常是高频小额收款及平台结算，异常大额整进整出需要重点说明。'},
  {key:'ecommerce',name:'电商/网络销售',pattern:/电商|电子商务|网络销售|网上销售|直播|平台销售/,topShare:.38,cv:.68,fast:.62,round:.42,netMargin:.10,feature:'平台回款和广告、物流支出应具有连续性，需区分平台结算与个人互转。'},
  {key:'manufacturing',name:'生产制造/加工',pattern:/制造|生产|加工|五金|机械|陶瓷|铝|钢|塑料|电子|家具|设备|材料/,topShare:.52,cv:.78,fast:.68,round:.52,netMargin:.12,feature:'应能看到相对稳定的销售回款和原料、工资、物流等支出，关注关联方走账及资金空转。'},
  {key:'trade',name:'批发贸易/经销',pattern:/贸易|批发|经销|销售|供应链|商贸|进出口|代理/,topShare:.56,cv:.82,fast:.76,round:.58,netMargin:.08,feature:'资金周转快、净流入较薄较常见，重点核验上下游真实性、同日快进快出及客户集中度。'},
  {key:'service',name:'服务/技术/租赁',pattern:/服务|咨询|技术|软件|信息|物流|运输|租赁|维修|设计|广告|培训/,topShare:.58,cv:.88,fast:.62,round:.52,netMargin:.18,feature:'通常原料支出较少，应关注收入来源、人工成本以及合同与回款的一致性。'},
  {key:'general',name:'综合经营/未明确行业',pattern:/.*/,topShare:.50,cv:.75,fast:.65,round:.50,netMargin:.10,feature:'采用通用银行流水模型；补充实际经营描述后可提高判断针对性。'}
];
function splitNames(value){return String(value||'').split(/[,，\n\r;；]+/).map(x=>x.trim()).filter(x=>x.length>=2)}
function riskInputs(){return{borrowerType:borrowerType.value||'enterprise',subjectName:subjectName.value.trim(),scope:businessScope.value.trim(),description:businessDescription.value.trim(),related:splitNames(relatedParties.value),loanAmount:Number(loanAmount.value||0),loanTerm:Math.max(1,Number(loanTerm.value||12))}}
function inferIndustry(input){const description=input.description||'';const scope=input.scope||'';return INDUSTRY_MODELS.find(x=>x.key!=='general'&&x.pattern.test(description))||INDUSTRY_MODELS.find(x=>x.key!=='general'&&x.pattern.test(scope))||INDUSTRY_MODELS[INDUSTRY_MODELS.length-1]}
function riskAnalysis(c){
  const input=riskInputs(),model=inferIndustry(input),allTxns=c.txns||[],txns=allTxns.filter(t=>!isExcluded(t));
  const months={},counterparties={},days={};let totalIncome=0,totalExpense=0,financialVolume=0,relatedVolume=0,unknownIncome=0,roundIncome=0;
  const relatedNames=[input.subjectName,...input.related].filter(Boolean);
  for(const t of txns){const income=Number(t.income||0),expense=Number(t.expense||0),volume=income+expense,month=t.month||'未识别日期',date=t.date||'';totalIncome+=income;totalExpense+=expense;const m=months[month]??={month,income:0,expense:0,count:0};m.income+=income;m.expense+=expense;m.count++;const cp=t.counterparty||'未识别';const p=counterparties[cp]??={name:cp,income:0,expense:0,count:0};p.income+=income;p.expense+=expense;p.count++;if(date){const d=days[date]??={income:0,expense:0};d.income+=income;d.expense+=expense}if(isFinancialCounterparty(t)||t.default_excluded)financialVolume+=volume;if(relatedNames.some(n=>`${cp} ${t.summary||''}`.includes(n)))relatedVolume+=volume;if(/未识别|未知|^-$/.test(cp))unknownIncome+=income;if(income>=10000&&Math.abs(income/10000-Math.round(income/10000))<.000001)roundIncome+=income}
  const monthRows=Object.values(months).filter(x=>/^20\d{2}-\d{2}$/.test(x.month)).sort((a,b)=>a.month.localeCompare(b.month));const activeMonths=monthRows.length;const avgIncome=activeMonths?totalIncome/activeMonths:0,avgExpense=activeMonths?totalExpense/activeMonths:0,avgNet=avgIncome-avgExpense;const variance=activeMonths?monthRows.reduce((s,x)=>s+Math.pow(x.income-avgIncome,2),0)/activeMonths:0,cv=avgIncome?Math.sqrt(variance)/avgIncome:0;const monthSerial=m=>Number(m.slice(0,4))*12+Number(m.slice(5,7));const span=activeMonths?monthSerial(monthRows.at(-1).month)-monthSerial(monthRows[0].month)+1:0,gapMonths=Math.max(0,span-activeMonths);const negativeMonths=monthRows.filter(x=>x.income<x.expense).length,negativeRatio=activeMonths?negativeMonths/activeMonths:0;
  financialVolume=allTxns.filter(t=>isFinancialCounterparty(t)||t.default_excluded).reduce((s,t)=>s+Number(t.income||0)+Number(t.expense||0),0);
  const cpRows=Object.values(counterparties).sort((a,b)=>b.income-a.income),topCp=cpRows[0]||{name:'无',income:0},topShare=totalIncome?topCp.income/totalIncome:0;const maxIncome=Math.max(0,...txns.map(t=>Number(t.income||0))),maxShare=totalIncome?maxIncome/totalIncome:0;const fastVolume=Object.values(days).reduce((s,d)=>s+Math.min(d.income,d.expense),0),fastShare=totalIncome?fastVolume/totalIncome:0;const totalVolume=totalIncome+totalExpense,allVolume=allTxns.reduce((s,t)=>s+Number(t.income||0)+Number(t.expense||0),0),financialShare=allVolume?financialVolume/allVolume:0,relatedShare=totalVolume?relatedVolume/totalVolume:0,unknownShare=totalIncome?unknownIncome/totalIncome:0,roundShare=totalIncome?roundIncome/totalIncome:0;
  const findings=[];const add=(severity,item,observed,reference,reason,action)=>findings.push({severity,item,observed,reference,reason,action});
  if(!input.scope&&!input.description)add('中','行业信息不足','未填写经营范围或实际经营描述','建议至少填写一项','只能使用通用模型，行业适配度有限。','补充主营业务、上下游、客单价及结算周期。');
  if(activeMonths<6)add('高','流水覆盖期过短',`${activeMonths}个月`,`至少6个月，通常建议12个月`,'样本不足以判断经营稳定性和季节性。','补充更完整的连续流水。');else if(activeMonths<12||gapMonths>0)add('中','流水连续性',`${activeMonths}个有效月，缺口${gapMonths}个月`,`建议连续12个月`,'覆盖不足或月份缺失会降低收入稳定性判断可信度。','核对是否遗漏账户或月份。');
  if(cv>model.cv*1.5)add('高','月度收入波动',`${(cv*100).toFixed(1)}%`,`行业参考不高于${(model.cv*100).toFixed(0)}%`,'收入波动显著高于当前经营模型。','结合淡旺季、项目回款和合同逐月解释。');else if(cv>model.cv)add('中','月度收入波动',`${(cv*100).toFixed(1)}%`,`行业参考不高于${(model.cv*100).toFixed(0)}%`,'收入稳定性偏弱。','核验高低月份的业务凭证。');
  if(topShare>model.topShare+.20)add('高','第一大交易对手集中',`${topCp.name}占${(topShare*100).toFixed(1)}%`,`行业参考不高于${(model.topShare*100).toFixed(0)}%`,'对单一客户依赖度较高，回款中断可能影响偿债。','核验合同、发票及客户持续合作情况。');else if(topShare>model.topShare)add('中','第一大交易对手集中',`${topCp.name}占${(topShare*100).toFixed(1)}%`,`行业参考不高于${(model.topShare*100).toFixed(0)}%`,'客户集中度偏高。','说明主要客户结构及合作年限。');
  if(fastShare>model.fast+.18)add('高','同日快进快出',`估算占收入${(fastShare*100).toFixed(1)}%`,`行业参考不高于${(model.fast*100).toFixed(0)}%`,'大量资金当日流入后流出，可能是低留存周转或过桥走账。','穿透核验对应上下游、合同和资金用途。');else if(fastShare>model.fast)add('中','同日快进快出',`估算占收入${(fastShare*100).toFixed(1)}%`,`行业参考不高于${(model.fast*100).toFixed(0)}%`,'资金留存偏低。','抽查大额同日收支交易。');
  if(roundShare>Math.max(.70,model.round+.18))add('高','大额整额收入',`占收入${(roundShare*100).toFixed(1)}%`,`行业参考不高于${(model.round*100).toFixed(0)}%`,'整万元回款比例过高，需排除临时归集或人为补流水。','核对大额整额交易的订单、发票和回款依据。');else if(roundShare>model.round)add('中','大额整额收入',`占收入${(roundShare*100).toFixed(1)}%`,`行业参考不高于${(model.round*100).toFixed(0)}%`,'整额交易占比较高。','抽查前十大整额收入。');
  if(relatedShare>.40)add('高','本人/关联方资金往来',`占总收支${(relatedShare*100).toFixed(1)}%`,`参考不高于40%`,'经营流水可能被关联方互转放大。','剔除关联方后重新评估真实经营收入。');else if(relatedShare>.20)add('中','本人/关联方资金往来',`占总收支${(relatedShare*100).toFixed(1)}%`,`参考不高于20%`,'关联方往来占比较高。','核验关联关系和交易背景。');
  if(financialShare>.35)add('中','金融及非经营性流水',`占总收支${(financialShare*100).toFixed(1)}%`,`参考不高于35%`,'贷款、还款、理财或内部资金可能影响经营流水口径。','按当前剔除规则复核并单独列示。');if(unknownShare>.20)add('中','交易对手识别不足',`未识别收入占${(unknownShare*100).toFixed(1)}%`,`参考不高于20%`,'交易对象不足会降低业务真实性核验能力。','补充清晰版流水或Excel原件。');if(negativeRatio>.60)add('高','月度净流入持续为负',`${negativeMonths}/${activeMonths}个月支出大于收入`,`参考不高于60%`,'账户资金留存和偿债缓冲偏弱。','核验是否存在多账户分流及主要支出用途。');else if(negativeRatio>.40)add('中','月度净流入偏弱',`${negativeMonths}/${activeMonths}个月支出大于收入`,`参考不高于40%`,'经营现金流稳定性偏弱。','结合其他账户和应收账款说明。');if(maxShare>.35)add('中','单笔收入占比过高',`${(maxShare*100).toFixed(1)}%`,`参考不高于35%`,'整体流水可能依赖少数大额回款。','核验该笔交易的合同、发票和到账凭证。');
  const annualizedIncome=activeMonths?avgIncome*12:0,loanRatio=annualizedIncome?input.loanAmount/annualizedIncome:0,monthlyPrincipal=input.loanAmount/input.loanTerm,coverage=monthlyPrincipal?avgNet/monthlyPrincipal:0;if(input.loanAmount>0){if(loanRatio>1.2)add('高','贷款金额与流水规模',`贷款/年化收入${loanRatio.toFixed(2)}倍`,`参考不高于1.20倍`,'拟贷款金额明显高于流水承载规模。','下调额度或补充其他可靠收入来源。');else if(loanRatio>.70)add('中','贷款金额与流水规模',`贷款/年化收入${loanRatio.toFixed(2)}倍`,`参考不高于0.70倍`,'贷款额度相对经营流水偏高。','结合毛利、存量负债和用途进一步测算。');if(avgNet<=0||coverage<1)add('高','简化还款覆盖',`${coverage.toFixed(2)}倍`,`参考不低于1.00倍`,'按月均净流水测算，偿还本金的缓冲不足。','核验多账户资金、毛利和现有负债；该指标仅作筛查。');else if(coverage<1.5)add('中','简化还款覆盖',`${coverage.toFixed(2)}倍`,`建议不低于1.50倍`,'还款缓冲偏薄。','审慎评估期限和还款方式。')}
  if(!findings.length)add('低','未发现明显规则异常','主要指标在当前模型参考范围内','仍需人工核验','未触发设定的重点异常规则。','抽查大额交易并结合征信、财报和经营资料复核。');const weights={高:18,中:9,低:0},score=Math.min(100,findings.reduce((s,x)=>s+(weights[x.severity]||0),0)),level=score>=55?'风险偏高':score>=30?'中等风险':score>=15?'需关注':'较低风险',riskClass=score>=55?'high':score>=15?'medium':'low';
  const evidence=[];for(const t of allTxns){const income=Number(t.income||0),expense=Number(t.expense||0),cp=t.counterparty||'未识别',reasons=[];if(relatedNames.some(n=>`${cp} ${t.summary||''}`.includes(n)))reasons.push('本人/关联方');if(!isExcluded(t)&&income>=10000&&Math.abs(income/10000-Math.round(income/10000))<.000001)reasons.push('大额整额收入');if(isFinancialCounterparty(t)||t.default_excluded)reasons.push('金融/非经营');if(!isExcluded(t)&&/未识别|未知/.test(cp))reasons.push('交易对手未识别');if(reasons.length)evidence.push({reason:reasons.join('、'),date:t.date||'',account:t.account||'',counterparty:cp,summary:t.summary||'',income,expense,source:t.source||''})}evidence.sort((a,b)=>(b.income+b.expense)-(a.income+a.expense));
  return{input,model,score,level,riskClass,findings,evidence:evidence.slice(0,30),metrics:{activeMonths,gapMonths,totalIncome,totalExpense,avgIncome,avgExpense,avgNet,cv,topCp:topCp.name,topShare,fastShare,roundShare,relatedShare,financialShare,unknownShare,negativeMonths,annualizedIncome,loanRatio,monthlyPrincipal,coverage}};
}
function riskTable(c){const r=riskAnalysis(c),m=r.metrics;const findingRows=r.findings.map(x=>`<tr><td><span class="pill ${x.severity==='高'?'high':x.severity==='中'?'medium':'low'}">${x.severity}</span></td><td>${esc(x.item)}</td><td>${esc(x.observed)}</td><td>${esc(x.reference)}</td><td>${esc(x.reason)}</td><td>${esc(x.action)}</td></tr>`).join(''),evidenceRows=r.evidence.map(x=>`<tr><td>${esc(x.reason)}</td><td>${esc(x.date)}</td><td>${esc(x.account)}</td><td>${esc(x.counterparty)}</td><td>${esc(x.summary)}</td><td>${money(x.income)}</td><td>${money(x.expense)}</td><td>${esc(x.source)}</td></tr>`).join('');return `<div class="risk-cards"><div class="risk-card"><span>综合风险</span><strong class="risk-${r.riskClass}">${esc(r.level)}</strong></div><div class="risk-card"><span>规则评分</span><strong>${r.score}/100</strong></div><div class="risk-card"><span>识别经营模型</span><strong>${esc(r.model.name)}</strong></div><div class="risk-card"><span>有效流水覆盖</span><strong>${m.activeMonths}个月</strong></div></div><div class="risk-note"><strong>模型说明：</strong>${esc(r.model.feature)} 实际经营描述优先于营业执照经营范围；所有异常均为“需要核验”，不直接认定虚假或欺诈。</div><div class="table-wrap"><table><thead><tr><th>建模项目</th><th>内容</th></tr></thead><tbody><tr><td>客户类型</td><td>${r.input.borrowerType==='enterprise'?'企业':'个人经营'}</td></tr><tr><td>客户/企业</td><td>${esc(r.input.subjectName||'未填写')}</td></tr><tr><td>经营范围</td><td>${esc(r.input.scope||'未填写')}</td></tr><tr><td>实际经营描述</td><td>${esc(r.input.description||'未填写')}</td></tr><tr><td>行业模型</td><td>${esc(r.model.name)}</td></tr><tr><td>月均收入 / 支出 / 净流入</td><td>${money(m.avgIncome)} / ${money(m.avgExpense)} / ${money(m.avgNet)}</td></tr><tr><td>第一大交易对手</td><td>${esc(m.topCp)}，占收入 ${(m.topShare*100).toFixed(1)}%</td></tr><tr><td>拟贷款测算</td><td>${r.input.loanAmount?`金额 ${money(r.input.loanAmount)}，期限 ${r.input.loanTerm}个月，简化月本金 ${money(m.monthlyPrincipal)}，覆盖 ${m.coverage.toFixed(2)}倍`:'未填写贷款条件'}</td></tr></tbody></table><table><thead><tr><th>程度</th><th>检测项目</th><th>检测结果</th><th>参考条件</th><th>银行风控解释</th><th>建议核验</th></tr></thead><tbody>${findingRows}</tbody></table><h3>异常交易证据（最多30笔）</h3><table><thead><tr><th>触发原因</th><th>日期</th><th>账户</th><th>交易对象</th><th>摘要</th><th>收入</th><th>支出</th><th>来源</th></tr></thead><tbody>${evidenceRows||'<tr><td colspan="8">暂无需单独列示的证据交易</td></tr>'}</tbody></table></div>`}
function renderCounterparties(){if(!rawData){counterparties.innerHTML='<span class="empty">上传后显示交易对象。</span>';return}const map=new Map();for(const t of rawData.transactions||[]){const name=t.counterparty||'未识别';const item=map.get(name)||{name,income:0,count:0};item.income+=Number(t.income||0);item.count++;map.set(name,item);}const items=[...map.values()].sort((a,b)=>b.income-a.income).slice(0,80);counterparties.innerHTML=items.map(x=>`<button class="cp ${clickedExcludes.has(x.name)?'active':''}" data-name="${x.name.replace(/"/g,'&quot;')}">${x.name} · ${money(x.income)}</button>`).join('')||'<span class="empty">未识别到交易对象。</span>';counterparties.querySelectorAll('.cp').forEach(btn=>btn.addEventListener('click',()=>{const name=btn.dataset.name;if(clickedExcludes.has(name))clickedExcludes.delete(name);else clickedExcludes.add(name);renderAll();}));}
function renderMetrics(c){document.getElementById('mValid').textContent=money(c.totalIncome);document.getElementById('mValidSub').textContent=directionMode.value==='income'?'收入方向':'收入支出合并口径';document.getElementById('mAvg').textContent=money(c.avg);document.getElementById('mAvgSub').textContent=`${c.validMonthCount} 个有效月份`;document.getElementById('mCount').textContent=c.effective.length.toLocaleString('zh-CN');document.getElementById('mCountSub').textContent=`原始 ${c.txns.length} 笔`;document.getElementById('mExcluded').textContent=c.excluded.length.toLocaleString('zh-CN');document.getElementById('mExcludedSub').textContent=excludeList().length?`对象 ${excludeList().length} 个`:'未手动剔除对象';}
function monthlyTable(c){const rows=c.monthly.map(x=>`<tr><td>${x.month}</td><td>${x.count}</td><td>${money(x.income)}</td><td>${money(x.expense)}</td><td>${money(x.income-x.expense)}</td><td>${money(x.interest)}</td><td>${x.interestCount||0}</td></tr>`).join('');return `<div class="table-wrap"><table><thead><tr><th>月份</th><th>有效笔数</th><th>有效收入流水</th><th>有效支出流水</th><th>净流水</th><th>银行季度结息</th><th>结息笔数</th></tr></thead><tbody>${rows||'<tr><td colspan="7">暂无有效流水</td></tr>'}<tr class="summary-row"><td>合计</td><td>${c.effective.length}</td><td>${money(c.totalIncome)}</td><td>${money(c.totalExpense)}</td><td>${money(c.totalIncome-c.totalExpense)}</td><td>${money(c.interestTotal)}</td><td>${c.interestCount}</td></tr><tr class="summary-row"><td>月均</td><td></td><td>${money(c.avg)}</td><td>${money(c.validMonthCount?c.totalExpense/c.validMonthCount:0)}</td><td>${money(c.validMonthCount?(c.totalIncome-c.totalExpense)/c.validMonthCount:0)}</td><td>${money(c.monthly.length?c.interestTotal/c.monthly.length:0)}</td><td></td></tr></tbody></table></div>`}
function fileTransactions(file){const txns=(rawData&&rawData.transactions)||[];return txns.filter(t=>String(t.source||'').startsWith(file.filename));}
function fileStatsTable(){
  const files=(rawData&&rawData.files)||[];
  if(!files.length)return '<div class="empty">暂无文件统计。</div>';
  return files.map(file=>{
    const txns=fileTransactions(file);
    const c=calc(txns);
    const rows=c.monthly.map(x=>`<tr><td>${x.month}</td><td>${x.count}</td><td>${money(x.income)}</td><td>${money(x.expense)}</td><td>${money(x.income-x.expense)}</td><td>${money(x.interest)}</td><td>${x.interestCount||0}</td></tr>`).join('');
    const duplicateText=Number(file.duplicate_transaction_count||0)>0?` · 跨文件去重 ${file.duplicate_transaction_count} 笔`:'';
    return `<section class="file-block"><div class="file-title"><h3>${esc(file.filename)}</h3><span>${esc(file.source_mode)} · 账号 ${esc(fileAccountText(file))} · 识别 ${file.transaction_count} 笔${duplicateText} · 当前有效 ${c.effective.length} 笔</span></div><div class="debug" style="margin-bottom:10px">${esc(fileInfoText(file))}</div><div class="table-wrap"><table><thead><tr><th>月份</th><th>有效笔数</th><th>有效收入流水</th><th>有效支出流水</th><th>净流水</th><th>银行季度结息</th><th>结息笔数</th></tr></thead><tbody>${rows||'<tr><td colspan="7">暂无有效流水</td></tr>'}<tr class="summary-row"><td>本文件合计</td><td>${c.effective.length}</td><td>${money(c.totalIncome)}</td><td>${money(c.totalExpense)}</td><td>${money(c.totalIncome-c.totalExpense)}</td><td>${money(c.interestTotal)}</td><td>${c.interestCount}</td></tr><tr class="summary-row"><td>本文件月均</td><td></td><td>${money(c.avg)}</td><td>${money(c.validMonthCount?c.totalExpense/c.validMonthCount:0)}</td><td>${money(c.validMonthCount?(c.totalIncome-c.totalExpense)/c.validMonthCount:0)}</td><td>${money(c.monthly.length?c.interestTotal/c.monthly.length:0)}</td><td></td></tr></tbody></table></div></section>`;
  }).join('');
}
function accountTable(c){
  if(!c.accountGroups.length)return '<div class="empty">暂无账户汇总。</div>';
  const summaryRows=c.accountGroups.map(account=>`<tr><td>${esc(account.account)}</td><td>${account.fileCount}</td><td>${account.monthCount}</td><td>${account.count}</td><td>${money(account.income)}</td><td>${money(account.expense)}</td><td>${money(account.income-account.expense)}</td><td>${money(account.monthCount?account.income/account.monthCount:0)}</td><td>${money(account.interest)}</td><td>${account.interestCount}</td></tr>`).join('');
  const totalTable=`<div class="table-wrap"><table><thead><tr><th>账号</th><th>文件数</th><th>月份数</th><th>有效笔数</th><th>有效收入流水</th><th>有效支出流水</th><th>净流水</th><th>账户月均收入</th><th>银行季度结息</th><th>结息笔数</th></tr></thead><tbody>${summaryRows}<tr class="summary-row"><td>全部账号合计</td><td>${c.accountGroups.reduce((s,x)=>s+Number(x.fileCount||0),0)}</td><td>${c.validMonthCount}</td><td>${c.effective.length}</td><td>${money(c.totalIncome)}</td><td>${money(c.totalExpense)}</td><td>${money(c.totalIncome-c.totalExpense)}</td><td>${money(c.validMonthCount?c.totalIncome/c.validMonthCount:0)}</td><td>${money(c.interestTotal)}</td><td>${c.interestCount}</td></tr></tbody></table></div>`;
  const sections=c.accountGroups.map(account=>{
    const files=(c.accountFiles||[]).filter(x=>x.account===account.account);
    const months=(c.accountMonthly||[]).filter(x=>x.account===account.account);
    const fileRows=files.map(x=>`<tr><td>${esc(x.file)}</td><td>${x.monthCount}</td><td>${x.count}</td><td>${money(x.income)}</td><td>${money(x.expense)}</td><td>${money(x.income-x.expense)}</td><td>${money(x.interest)}</td><td>${x.interestCount}</td></tr>`).join('');
    const monthRows=months.map(x=>`<tr><td>${x.month}</td><td>${x.count}</td><td>${money(x.income)}</td><td>${money(x.expense)}</td><td>${money(x.income-x.expense)}</td><td>${money(x.interest)}</td><td>${x.interestCount}</td></tr>`).join('');
    return `<section class="file-block"><div class="file-title"><h3>账号：${esc(account.account)}</h3><span>${account.fileCount} 个文件 · ${account.monthCount} 个月 · ${account.count} 笔有效流水</span></div><div class="table-wrap"><table><thead><tr><th>有效收入流水</th><th>有效支出流水</th><th>净流水</th><th>账户月均收入</th><th>银行季度结息</th><th>结息笔数</th></tr></thead><tbody><tr class="summary-row"><td>${money(account.income)}</td><td>${money(account.expense)}</td><td>${money(account.income-account.expense)}</td><td>${money(account.monthCount?account.income/account.monthCount:0)}</td><td>${money(account.interest)}</td><td>${account.interestCount}</td></tr></tbody></table><table><thead><tr><th>该账号包含文件</th><th>月份数</th><th>有效笔数</th><th>有效收入流水</th><th>有效支出流水</th><th>净流水</th><th>银行季度结息</th><th>结息笔数</th></tr></thead><tbody>${fileRows||'<tr><td colspan="8">暂无文件明细</td></tr>'}</tbody></table><table><thead><tr><th>月份</th><th>有效笔数</th><th>有效收入流水</th><th>有效支出流水</th><th>净流水</th><th>银行季度结息</th><th>结息笔数</th></tr></thead><tbody>${monthRows||'<tr><td colspan="7">暂无月度明细</td></tr>'}</tbody></table></div></section>`;
  }).join('');
  return totalTable+sections;
}
function detailTable(c){const rows=c.txns.slice().sort((a,b)=>(a.date||'').localeCompare(b.date||'')).map(t=>{const off=!isEffective(t);return `<tr><td>${t.date||''}</td><td>${t.account||''}</td><td>${t.counterparty||''}</td><td>${t.summary||''}</td><td>${money(t.income)}</td><td>${money(t.expense)}</td><td>${money(t.other_amount)}</td><td>${t.balance==null?'':money(t.balance)}</td><td>${t.source||''}</td><td>${off?'<span class="pill off">剔除</span>':'<span class="pill">有效</span>'}</td></tr>`}).join('');return `<div class="table-wrap"><table><thead><tr><th>日期</th><th>账户</th><th>交易对象</th><th>摘要/备注</th><th>收入</th><th>支出</th><th>其他</th><th>余额</th><th>来源</th><th>状态</th></tr></thead><tbody>${rows||'<tr><td colspan="10">暂无交易明细</td></tr>'}</tbody></table></div>`}
function summaryTable(c){const files=(rawData.files||[]).map(f=>`<tr><td>${esc(f.filename)}</td><td>${esc(fileAccountText(f))}</td><td>${esc(fileInfoText(f))}</td><td>${esc(f.source_mode)}</td><td>${f.transaction_count}</td><td>${f.duplicate_transaction_count||0}</td><td>${f.unique_transaction_count??f.transaction_count}</td><td>${f.text_length||0}</td></tr>`).join('');return `<div class="table-wrap"><table><thead><tr><th>项目</th><th>数值</th></tr></thead><tbody><tr><td>有效流水合计</td><td>${money(c.totalIncome)}</td></tr><tr><td>月均有效流水</td><td>${money(c.avg)}</td></tr><tr><td>同账户合并数量</td><td>${c.accountGroups.length} 个账户组</td></tr><tr><td>跨文件重复交易</td><td>${rawData.duplicate_transaction_count||0} 笔</td></tr><tr><td>银行季度结息合计</td><td>${money(c.interestTotal)}，${c.interestCount} 笔</td></tr><tr><td>有效笔数</td><td>${c.effective.length}</td></tr><tr><td>剔除笔数</td><td>${c.excluded.length}</td></tr><tr><td>统计月份数</td><td>${c.monthly.length}</td></tr><tr><td>剔除对象</td><td>${esc(excludeList().join('，')||'无')}</td></tr></tbody></table><table><thead><tr><th>文件</th><th>识别账号</th><th>文件信息</th><th>读取方式</th><th>识别笔数</th><th>跨文件重复</th><th>去重后笔数</th><th>文本长度</th></tr></thead><tbody>${files}</tbody></table></div>`}
function exportPayload(){
  if(!rawData)throw new Error('请先完成统计后再下载。');
  const c=calc();
  const filters=`${directionMode.value==='income'?'收入流水':'收入和支出'}；${excludeDefault.checked?'剔除常见非经营流水':'不自动剔除常见非经营流水'}；${excludeFinancial.checked?'剔除银行/金融机构交易对手':'不自动剔除银行/金融机构交易对手'}；剔除对象：${excludeList().join('、')||'无'}`;
  const blocks=[
    {title:'汇总栏',headers:['项目','数值'],rows:[
      ['有效流水合计',c.totalIncome],['月均有效流水',c.avg],['有效支出流水',c.totalExpense],['净流水',c.totalIncome-c.totalExpense],['银行季度结息',c.interestTotal],['结息笔数',c.interestCount],['有效笔数',c.effective.length],['剔除笔数',c.excluded.length],['统计月份数',c.validMonthCount],['同账户合并数量',c.accountGroups.length],['筛选口径',filters]
    ]},
    {title:'月度汇总',headers:['月份','有效笔数','有效收入流水','有效支出流水','净流水','银行季度结息','结息笔数'],rows:[
      ...c.monthly.map(x=>[x.month,x.count,x.income,x.expense,x.income-x.expense,x.interest,x.interestCount||0]),
      ['合计',c.effective.length,c.totalIncome,c.totalExpense,c.totalIncome-c.totalExpense,c.interestTotal,c.interestCount],
      ['月均','',c.avg,c.validMonthCount?c.totalExpense/c.validMonthCount:0,c.validMonthCount?(c.totalIncome-c.totalExpense)/c.validMonthCount:0,c.monthly.length?c.interestTotal/c.monthly.length:0,'']
    ]},
  ];
  const risk=riskAnalysis(c),rm=risk.metrics;
  blocks.push({title:'银行风控概览',headers:['项目','结果'],rows:[['综合风险',risk.level],['规则评分',risk.score],['客户类型',risk.input.borrowerType==='enterprise'?'企业':'个人经营'],['客户/企业',risk.input.subjectName||'未填写'],['经营范围',risk.input.scope||'未填写'],['实际经营描述',risk.input.description||'未填写'],['识别行业模型',risk.model.name],['模型说明',risk.model.feature],['有效月份',rm.activeMonths],['缺口月份',rm.gapMonths],['月均收入',rm.avgIncome],['月均支出',rm.avgExpense],['月均净流入',rm.avgNet],['月度收入波动率',`${(rm.cv*100).toFixed(1)}%`],['第一大交易对手',rm.topCp],['第一大对手收入占比',`${(rm.topShare*100).toFixed(1)}%`],['同日快进快出估算占比',`${(rm.fastShare*100).toFixed(1)}%`],['大额整额收入占比',`${(rm.roundShare*100).toFixed(1)}%`],['关联方往来占比',`${(rm.relatedShare*100).toFixed(1)}%`],['金融及非经营流水占比',`${(rm.financialShare*100).toFixed(1)}%`],['拟贷款金额',risk.input.loanAmount||'未填写'],['贷款期限（月）',risk.input.loanTerm],['简化月本金',rm.monthlyPrincipal],['简化还款覆盖',risk.input.loanAmount?`${rm.coverage.toFixed(2)}倍`:'未测算'],['重要说明','异常表示需要核验，不直接认定虚假或欺诈；结果不等同于银行最终审批。']]});
  blocks.push({title:'银行风控异常检测',headers:['程度','检测项目','检测结果','参考条件','银行风控解释','建议核验'],rows:risk.findings.map(x=>[x.severity,x.item,x.observed,x.reference,x.reason,x.action])});
  blocks.push({title:'风控证据交易',headers:['触发原因','日期','账户','交易对象','摘要','收入','支出','来源'],rows:risk.evidence.map(x=>[x.reason,x.date,x.account,x.counterparty,x.summary,x.income,x.expense,x.source])});
  const fileRows=[];
  for(const file of rawData.files||[]){
    const fc=calc(fileTransactions(file));
    for(const x of fc.monthly){
      fileRows.push([file.filename,file.source_mode,fileInfoText(file),x.month,x.count,x.income,x.expense,x.income-x.expense,x.interest,x.interestCount||0]);
    }
    fileRows.push([file.filename,'本文件合计',fileInfoText(file),'合计',fc.effective.length,fc.totalIncome,fc.totalExpense,fc.totalIncome-fc.totalExpense,fc.interestTotal,fc.interestCount]);
  }
  blocks.push({title:'单文件统计',headers:['文件','读取方式','账户信息','月份','有效笔数','有效收入流水','有效支出流水','净流水','银行季度结息','结息笔数'],rows:fileRows});
  blocks.push({title:'同账户汇总',headers:['账号','文件数','月份数','有效笔数','有效收入流水','有效支出流水','净流水','账户月均收入','银行季度结息','结息笔数'],rows:[
    ...c.accountGroups.map(x=>[x.account,x.fileCount,x.monthCount,x.count,x.income,x.expense,x.income-x.expense,x.monthCount?x.income/x.monthCount:0,x.interest,x.interestCount]),
    ['全部账号合计',c.accountGroups.reduce((s,x)=>s+Number(x.fileCount||0),0),c.validMonthCount,c.effective.length,c.totalIncome,c.totalExpense,c.totalIncome-c.totalExpense,c.validMonthCount?c.totalIncome/c.validMonthCount:0,c.interestTotal,c.interestCount]
  ]});
  blocks.push({title:'账户文件明细',headers:['账户','文件','月份数','有效笔数','有效收入流水','有效支出流水','净流水','银行季度结息','结息笔数'],rows:(c.accountFiles||[]).map(x=>[x.account,x.file,x.monthCount,x.count,x.income,x.expense,x.income-x.expense,x.interest,x.interestCount])});
  blocks.push({title:'账户月度明细',headers:['账户','月份','有效笔数','有效收入流水','有效支出流水','净流水','银行季度结息','结息笔数'],rows:c.accountMonthly.map(x=>[x.account,x.month,x.count,x.income,x.expense,x.income-x.expense,x.interest,x.interestCount])});
  blocks.push({title:'交易明细',headers:['日期','账户','交易对象','摘要/备注','收入','支出','其他','余额','来源','状态','类别'],rows:c.txns.slice().sort((a,b)=>(a.date||'').localeCompare(b.date||'')).map(t=>[t.date||'',t.account||'',t.counterparty||'',t.summary||'',Number(t.income||0),Number(t.expense||0),Number(t.other_amount||0),t.balance==null?'':Number(t.balance||0),t.source||'',isEffective(t)?'有效':'剔除',t.category||''])});
  blocks.push({title:'文件信息',headers:['文件','识别账号','账户信息','读取方式','识别笔数','跨文件重复','去重后笔数','文本长度'],rows:(rawData.files||[]).map(f=>[f.filename,fileAccountText(f),fileInfoText(f),f.source_mode,f.transaction_count,f.duplicate_transaction_count||0,f.unique_transaction_count??f.transaction_count,f.text_length||0])});
  return{title:'流水统计报告',generated_at:new Date().toLocaleString('zh-CN',{hour12:false}),filters_text:filters,blocks};
}
function filenameFromDisposition(header,fallback){if(!header)return fallback;const m=header.match(/filename\*=UTF-8''([^;]+)/i);if(m){try{return decodeURIComponent(m[1])}catch(_){return fallback}}const n=header.match(/filename="?([^";]+)"?/i);return n?n[1]:fallback;}
async function downloadReport(format){
  if(!rawData){statusEl.textContent='请先完成统计后再下载。';return}
  statusEl.textContent='正在生成下载文件...';
  try{
    const res=await fetch('/flow/export',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({format,data:exportPayload()})});
    if(!res.ok){let msg='生成下载文件失败';try{const err=await res.json();msg=err.error||msg}catch(_){msg=await res.text()||msg}throw new Error(msg)}
    const blob=await res.blob();
    const ext=format==='pdf'?'pdf':'xlsx';
    const name=filenameFromDisposition(res.headers.get('Content-Disposition'),`流水统计报告.${ext}`);
    const url=URL.createObjectURL(blob);
    const a=document.createElement('a');
    a.href=url;a.download=name;document.body.appendChild(a);a.click();a.remove();URL.revokeObjectURL(url);
    statusEl.textContent=`已生成下载文件：${name}`;
  }catch(err){statusEl.textContent=err.message;debug.textContent=err.stack||String(err)}
}
function renderResult(){if(!rawData){return}const c=calc();renderMetrics(c);if(activeTab==='monthly')result.innerHTML=monthlyTable(c);if(activeTab==='risk')result.innerHTML=riskTable(c);if(activeTab==='files')result.innerHTML=fileStatsTable();if(activeTab==='account')result.innerHTML=accountTable(c);if(activeTab==='details')result.innerHTML=detailTable(c);if(activeTab==='summary')result.innerHTML=summaryTable(c);debug.textContent=`文件数量：${(rawData.files||[]).length}\n原始交易：${c.txns.length} 笔\n有效交易：${c.effective.length} 笔\n同账户合并：${c.accountGroups.length} 个账户组\n剔除交易：${c.excluded.length} 笔\n银行季度结息：${money(c.interestTotal)}，${c.interestCount} 笔\n默认剔除关键词：${(rawData.default_exclude_keywords||[]).join('，')}`;}
function renderAll(){renderCounterparties();renderResult();}
['dragenter','dragover'].forEach(e=>drop.addEventListener(e,ev=>{ev.preventDefault();drop.classList.add('drag')}));['dragleave','drop'].forEach(e=>drop.addEventListener(e,ev=>{ev.preventDefault();drop.classList.remove('drag')}));drop.addEventListener('drop',ev=>{fileInput.files=ev.dataTransfer.files;updateFile();});fileInput.addEventListener('change',updateFile);
async function waitJob(jobId){
  while(true){
    const res=await fetch(`/flow/progress?job_id=${encodeURIComponent(jobId)}`);
    const job=await res.json();
    if(!res.ok)throw new Error(job.error||'无法读取进度');
    setProgress(job.percent||0);
    statusEl.textContent=`${job.percent||0}%  ${job.message||'处理中'}`;
    if(job.status==='needs_password')return{needsPassword:true,job};
    if(job.status==='cancelled')return{cancelled:true,job};
    if(job.status==='cancelling'){await new Promise(r=>setTimeout(r,500));continue;}
    if(job.status==='done')return{result:job.result};
    if(job.status==='error')throw new Error((job.error||job.message||'统计失败')+(job.trace?`\n${job.trace}`:''));
    await new Promise(r=>setTimeout(r,900));
  }
}
async function handleJobResult(done){
  if(done.cancelled){
    setRunning(false);
    resumeBtn.style.display='none';
    statusEl.textContent=done.job.message||'已停止统计。';
    debug.textContent='任务已停止，没有生成新的统计结果。';
    return;
  }
  if(done.needsPassword){
    setRunning(false);
    const files=done.job.password_files&&done.job.password_files.length?done.job.password_files:[done.job.password_file||'该文件'];
    const finished=done.job.result&&done.job.result.transactions?done.job.result.transactions.length:0;
    statusEl.textContent=`${files.length} 个文件需要密码：${files.join('，')}。\n请在“文件密码”输入框填写后，点击“输入密码后继续”。已完成的文件不会重复识别。`;
    debug.textContent=`等待密码的文件：\n${files.join('\n')}\n\n已先完成识别：${finished} 笔流水。`;
    resumeBtn.style.display='inline-flex';
    passwordsInput.focus();
    return;
  }
  rawData=done.result;
  setRunning(false);
  resumeBtn.style.display='none';
  setDownloadEnabled(true);
  statusEl.textContent=`统计完成：识别 ${(rawData.transactions||[]).length} 笔流水。`;
  setProgress(100);
  renderAll();
}
form.addEventListener('submit',async e=>{
  e.preventDefault();
  if(!fileInput.files.length){statusEl.textContent='请先选择流水文件。';return}
  const fd=new FormData(form);
  setRunning(true);
  setDownloadEnabled(false);
  resumeBtn.style.display='none';
  setProgress(0);
  result.innerHTML='<div class="empty">正在读取流水，请稍候。</div>';
  try{
    const start=await fetch('/flow/analyze/start',{method:'POST',body:fd});
    const started=await start.json();
    if(!start.ok)throw new Error(started.error||'提交任务失败');
    currentJobId=started.job_id;
    await handleJobResult(await waitJob(currentJobId));
  }catch(err){
    statusEl.textContent=err.message;
    debug.textContent=err.stack||String(err);
  }finally{
    if(stopBtn.style.display!=='none')setRunning(false);
  }
});
resumeBtn.addEventListener('click',async()=>{
  if(!currentJobId){statusEl.textContent='没有可继续的任务，请重新上传。';return}
  const fd=new FormData();
  fd.append('job_id',currentJobId);
  fd.append('passwords',passwordsInput.value||'');
  resumeBtn.disabled=true;
  setRunning(true);
  statusEl.textContent='已收到密码，正在继续统计...';
  try{
    const start=await fetch('/flow/analyze/resume',{method:'POST',body:fd});
    const started=await start.json();
    if(!start.ok)throw new Error(started.error||'继续任务失败');
    await handleJobResult(await waitJob(currentJobId));
  }catch(err){
    statusEl.textContent=err.message;
    debug.textContent=err.stack||String(err);
  }finally{
    resumeBtn.disabled=false;
    if(stopBtn.style.display!=='none')setRunning(false);
  }
});
stopBtn.addEventListener('click',async()=>{
  if(!currentJobId){statusEl.textContent='当前没有正在统计的任务。';return}
  const fd=new FormData();
  fd.append('job_id',currentJobId);
  stopBtn.disabled=true;
  statusEl.textContent='正在停止任务...';
  try{
    const res=await fetch('/flow/analyze/cancel',{method:'POST',body:fd});
    const data=await res.json();
    if(!res.ok)throw new Error(data.error||'停止任务失败');
  }catch(err){
    stopBtn.disabled=false;
    statusEl.textContent=err.message;
  }
});
riskAnalyzeBtn.addEventListener('click',()=>{const tab=document.querySelector('[data-tab="risk"]');if(tab)tab.click();else{activeTab='risk';renderResult()}});
[borrowerType,subjectName,businessScope,businessDescription,relatedParties,loanAmount,loanTerm].forEach(el=>{el.addEventListener('input',()=>{if(rawData&&activeTab==='risk')renderResult()});el.addEventListener('change',()=>{if(rawData&&activeTab==='risk')renderResult()})});
document.querySelectorAll('.tab').forEach(btn=>btn.addEventListener('click',()=>{document.querySelectorAll('.tab').forEach(x=>x.classList.remove('active'));btn.classList.add('active');activeTab=btn.dataset.tab;renderResult();}));document.getElementById('applyFilter').addEventListener('click',renderAll);excludeNames.addEventListener('input',renderAll);excludeDefault.addEventListener('change',renderAll);excludeFinancial.addEventListener('change',renderAll);directionMode.addEventListener('change',renderAll);document.getElementById('resetBtn').addEventListener('click',()=>{form.reset();rawData=null;currentJobId=null;setRunning(false);setDownloadEnabled(false);resumeBtn.style.display='none';clickedExcludes.clear();updateFile();setProgress(0);statusEl.textContent='请选择流水文件。';counterparties.innerHTML='<span class="empty">上传后显示交易对象。</span>';result.innerHTML='<div class="empty">上传流水后，这里会按月份列出有效流水，并在底部显示汇总。</div>';debug.textContent='等待统计。';});document.getElementById('downloadPdfBtn').addEventListener('click',()=>downloadReport('pdf'));document.getElementById('downloadExcelBtn').addEventListener('click',()=>downloadReport('excel'));document.getElementById('printBtn').addEventListener('click',()=>window.print());document.getElementById('copyBtn').addEventListener('click',async()=>{await navigator.clipboard.writeText(result.innerText);statusEl.textContent='当前汇总已复制。'});
document.getElementById('resetBtn').addEventListener('click',()=>{borrowerType.value='enterprise';subjectName.value='';businessScope.value='';businessDescription.value='';relatedParties.value='';loanAmount.value='';loanTerm.value='12'});
setDownloadEnabled(false);
</script></body></html>"""


class Handler(BaseHTTPRequestHandler):
    def send_json(self, data: dict, status: int = 200) -> None:
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/flow/progress":
            job_id = (parse_qs(parsed.query).get("job_id") or [""])[0]
            job = get_job(job_id)
            self.send_json(job or {"error": "任务不存在"}, 200 if job else 404)
            return
        if parsed.path in {"/", "/flow"}:
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(INDEX.encode("utf-8"))
            return
        self.send_error(404)

    def do_POST(self) -> None:
        if self.path == "/flow/export":
            try:
                length = int(self.headers.get("Content-Length") or "0")
                payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                content, filename, content_type = build_flow_export(payload.get("data") or {}, payload.get("format") or "")
                self.send_response(200)
                self.send_header("Content-Type", content_type)
                self.send_header("Content-Length", str(len(content)))
                self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(filename)}")
                self.end_headers()
                self.wfile.write(content)
            except Exception as exc:
                self.send_json({"error": str(exc), "trace": traceback.format_exc()}, 500)
            return
        if self.path not in {"/flow/analyze/start", "/flow/analyze/resume", "/flow/analyze/cancel"}:
            self.send_error(404)
            return
        try:
            form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST"})
            if self.path == "/flow/analyze/cancel":
                job_id = normalize_text(form.getfirst("job_id", ""))
                if not job_id:
                    raise ValueError("缺少任务编号。")
                job = get_job(job_id)
                if not job:
                    raise ValueError("任务不存在或已经结束。")
                if job.get("status") in {"done", "error", "cancelled"}:
                    self.send_json({"job_id": job_id, "status": job.get("status"), "message": job.get("message", "")})
                    return
                set_job(job_id, status="cancelling", cancel_requested=True, message="正在停止任务...")
                self.send_json({"job_id": job_id, "status": "cancelling", "message": "正在停止任务..."})
                return
            if self.path == "/flow/analyze/resume":
                job_id = normalize_text(form.getfirst("job_id", ""))
                if not job_id:
                    raise ValueError("缺少任务编号，请重新上传文件。")
                job = get_job(job_id)
                pending_files = job.get("pending_password_files") or job.get("files", [])
                files = [Path(item) for item in pending_files]
                if not files:
                    raise ValueError("没有需要继续处理的文件，请重新上传。")
                passwords = parse_passwords(form.getfirst("passwords", ""))
                set_job(
                    job_id,
                    status="queued",
                    percent=get_job(job_id).get("percent", 0),
                    message=f"已收到密码，只继续处理 {len(files)} 个缺密码文件",
                    total_files=len(files),
                    current_file=0,
                    cancel_requested=False,
                )
                thread = threading.Thread(target=run_flow_job, args=(job_id, files, passwords, True), daemon=True)
                thread.start()
                self.send_json({"job_id": job_id, "status": "queued"})
                return

            if "file" not in form:
                raise ValueError("未收到文件。")
            UPLOADS.mkdir(parents=True, exist_ok=True)
            items = form["file"] if isinstance(form["file"], list) else [form["file"]]
            files = []
            for idx, item in enumerate(items, start=1):
                if not getattr(item, "filename", None):
                    continue
                filename = Path(item.filename or f"upload-{idx}").name
                path = UPLOADS / filename
                if path.exists():
                    path = UPLOADS / f"{path.stem}-{idx}{path.suffix}"
                path.write_bytes(item.file.read())
                files.append(path)
            if not files:
                raise ValueError("未收到可统计的文件。")
            passwords = parse_passwords(form.getfirst("passwords", ""))
            job_id = uuid.uuid4().hex
            set_job(job_id, status="queued", percent=0, message="任务已提交", total_files=len(files), current_file=0, files=[str(path) for path in files], cancel_requested=False)
            thread = threading.Thread(target=run_flow_job, args=(job_id, files, passwords), daemon=True)
            thread.start()
            self.send_json({"job_id": job_id, "status": "queued"})
        except Exception as exc:
            self.send_json({"error": str(exc), "trace": traceback.format_exc()}, 500)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8790"))
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"本地流水统计平台：http://127.0.0.1:{port}/flow")
    server.serve_forever()
